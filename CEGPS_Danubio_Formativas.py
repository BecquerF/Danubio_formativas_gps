import logging
import io
import base64
import textwrap
import html as html_module
import tempfile
from pathlib import Path
import dash
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import plotly.io as pio

# Selenium and WebDriver Manager imports
import os
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager

# Set up Chrome options
options = Options()
if os.path.exists("/usr/bin/chromium"):
    options.binary_location = "/usr/bin/chromium"
elif os.path.exists("/usr/bin/google-chrome"):
    options.binary_location = "/usr/bin/google-chrome"
    
options.add_argument("--headless=new")  # Use the new headless mode
options.add_argument("--no-sandbox")     # Required for running headless in certain environments
options.add_argument("--disable-dev-shm-usage")  # Overcome limited resource problems

# Set up the ChromeDriver service
service = Service(ChromeDriverManager().install())

# Initialize the Chrome WebDriver
driver = webdriver.Chrome(service=service, options=options)

# Navigate to the Python website
driver.get("https://www.python.org")

# Print the title of the page
print(driver.title)

# Close the browser
driver.quit()

logging.basicConfig(level=logging.INFO)

import logging
import pandas as pd
from openpyxl.drawing.image import Image as ExcelImage
from datetime import datetime
from pathlib import Path

# WeasyPrint setup
try:
    from weasyprint import HTML as WeasyHTML
except Exception as e:
    logging.warning("WeasyPrint no está disponible: %s", e)
    WeasyHTML = None

# Dash imports
from dash import Dash, dcc, html, dash_table, Input, Output, State, no_update, ctx, ALL
import dash_auth

# ReportLab imports
try:
    from PIL import Image as PILImage
except ImportError:
    PILImage = None

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import inch
from reportlab.lib.utils import ImageReader
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Table, TableStyle
from reportlab.lib import colors
import time

# Base directory and font directory
BASE_DIR = Path(__file__).resolve().parent
FONT_DIR = BASE_DIR / "assets" / "fonts"

def register_pdf_fonts():
    try:
        pdfmetrics.registerFont(TTFont("ClashDisplay-Semibold", str(FONT_DIR / "ClashDisplay-Semibold.otf")))
    except Exception:
        pass
    try:
        pdfmetrics.registerFont(TTFont("Manrope-Light", str(FONT_DIR / "Manrope-Light.ttf")))
    except Exception:
        pass

# Register fonts
register_pdf_fonts()
logging.basicConfig(level=logging.INFO)

# Load data from Excel
df = pd.read_excel("GPS_Formativas_2026.xlsx")

# Convert Date column to datetime
# Supongamos que tu columna se llama 'Date'
df["Date"] = pd.to_datetime(
    df["Date"],
    dayfirst=True,       # porque tus fechas son dd/mm/yy
    errors="coerce"      # valores inválidos se convierten en NaT
)

# Process Duration column
if "Duration" in df.columns:
    if pd.api.types.is_numeric_dtype(df["Duration"]):
        df["Duration"] = pd.to_timedelta(df["Duration"], unit="D", errors="coerce").dt.total_seconds() / 60.0
    else:
        df["Duration"] = pd.to_timedelta(df["Duration"], errors="coerce").dt.total_seconds() / 60.0


# Date calculations
fecha_max = df["Date"].max()
ultimos21 = fecha_max - pd.Timedelta(days=21)
ultimos7 = fecha_max - pd.Timedelta(days=7)

# Drop unnecessary columns
columnas_eliminar = [
    "Period Number",
    "Work/Rest Ratio",
    "Max Heart Rate",
    "Avg Heart Rate",
    "Max HR (% Max)",
    "Avg HR (% Max)",
    "HR Exertion",
    "Red Zone",
    "Heart Rate Band 1 Duration",
    "Heart Rate Band 2 Duration",
    "Heart Rate Band 3 Duration",
    "Heart Rate Band 5 Duration",
    "Heart Rate Band 6 Duration",
    "Energy",
    "High Metabolic Load Distance",
    "Standing Distance",
    "Walking Distance",
    "Jogging Distance",
    "Running Distance",
    "Athlete Participation Tags"
]

df = df.drop(columns=columnas_eliminar, errors="ignore")

app = Dash(__name__)
app.config.suppress_callback_exceptions = True
server = app.server

# Set secret key for session management
SECRET_KEY = os.environ["SECRET_KEY"]
server.secret_key = SECRET_KEY

# Define valid username and password pairs for authentication
VALID_USERNAME_PASSWORD_PAIRS = {
    "Danubioformativas": "formativas2026"
}

# Setup basic authentication if dash_auth is available
if dash_auth is not None:
    auth = dash_auth.BasicAuth(app, VALID_USERNAME_PASSWORD_PAIRS)
else:
    auth = None

app.title = "DATA LOAD - Sports Performance Platform"

# Define the metrics related to sports performance
metricas = [
    "Distance",
    "Meterage Per Minute",
    "Player Load",
    "Player Load Per Minute",
    "Max Velocity",
    "Accel + Decel Efforts",
    "Accel + Decel Efforts Per Minute",
    "High Speed Distance",
    "High Speed Distance Per Minute",
    "High Speed Efforts",
    "Sprint Distance",
    "Sprint Dist Per Min",
    "Sprint Efforts",
    "Impacts"
]

# Define radar metrics for specific visualizations
metricas_radar = [
    "Meterage Per Minute",
    "Accel + Decel Efforts Per Minute",
    "High Speed Distance Per Minute",
    "Sprint Dist Per Min",
    "High Speed Efforts",
    "Sprint Efforts"
]

# Combine radar metrics with additional metrics for averages
metricas_promedios = metricas_radar + ["Max Velocity", "Duration"]

# ======================================================
# ACTIVIDAD COMPARATIVA INDIVIDUAL
# ======================================================

# Copy base metrics
metricas_base = metricas.copy()

# Calculate average metrics per player
df_promedios = (
    df.groupby("Player Name")[metricas_base]
    .mean()
    .reset_index()
)

# Rename columns to indicate they are averages
df_promedios.rename(
    columns={col: f"{col} Prom" for col in metricas_base},
    inplace=True
)

# Calculate cumulative metrics per player
df_acumulados = (
    df.groupby("Player Name")[metricas_base]
    .sum()
    .reset_index()
)

df_Actividad_Comparativa_Individual = pd.merge(
    df_acumulados,
    df_promedios,
    on="Player Name"
)

# Order columns
orden = ["Player Name"]
for m in metricas_base:
    orden.extend([m, f"{m} Prom"])

# Reorganize the DataFrame according to the specified order
df_Actividad_Comparativa_Individual = df_Actividad_Comparativa_Individual[orden]

# ======================================================
# Configure references and tab titles
referencias = [
    "Category",
    "Player Name",
    "Athlete Tags",
    "Game Tags",
    "Period Tags"
]

tab_titles = {
    "comparativas": "Comparativo",
    "cronologico": "Cronológico",
    "actividad": "Actividad_por_Jugador",
    "actividad_comparativa": "Actividad_Comparativa_Individual",
    "actividad_promedios": "Actividad_Promedios",
    "acwr": "ACWR_Zona_Segura",
    "plyr_vs_plyr": "PLYR_vs_PLYR",
    "informe": "Informe"
}

# Define sections for the report
informe_sections = [
    {"label": "Actividad", "value": "actividad"},
    {"label": "Actividad Comparativa Individual", "value": "actividad_comparativa"},
    {"label": "Actividad/Promedios", "value": "actividad_promedios"},
    {"label": "ACWR", "value": "acwr"},
    {"label": "PLYR vs PLYR", "value": "plyr_vs_plyr"},
    {"label": "Comparativo", "value": "comparativas"},
    {"label": "Cronológico", "value": "cronologico"}
]

# Handle logo image
LOGO_PATH = Path("assets/logo_dataload_2.png")
LOGO_BASE64 = ""

if LOGO_PATH.exists():
    with open(LOGO_PATH, "rb") as logo_file:
        LOGO_BASE64 = base64.b64encode(logo_file.read()).decode("ascii")

# ------------------------------------------------------
# Define utility functions
def summarize_items(items, max_items=3, default="todas"):
    if not items:
        return default
    labels = [str(item) for item in items if item is not None]

    if len(labels) == 0:
        return default
    if len(labels) <= max_items:
        return " / ".join(labels)

    return " / ".join(labels[:max_items]) + f" +{len(labels)-max_items}"

def build_chart_title(tab, categorias, metricas, referencia):
    # Generate summarized text for categories and metrics
    categoria_text = summarize_items(categorias, max_items=3)
    metrica_text = summarize_items(metricas, max_items=3)

    # Default title format
    title = tab_titles.get(tab, tab)  # Get default title based on tab

    # Construct titles based on the selected tab
    if tab == "comparativas":
        title = f"Comparativo de {metrica_text} por {referencia}"
    elif tab == "cronologico":
        title = f"Evolución cronológica de {metrica_text}"
    elif tab == "actividad":
        title = "Actividad por jugador"
    elif tab == "acwr":
        title = "ACWR - Últimos 7 días vs 21 días"
    elif tab == "actividad_comparativa":
        title = "Actividad comparativa individual"
    elif tab == "actividad_promedios":
        title = "Actividad / Promedios"
    elif tab == "plyr_vs_plyr":
        title = "Comparativa Jugador vs Jugador"

    # Add category info if available
    if categorias:
        title += f" - Categoría(s): {categoria_text}"

    return title

def build_download_metadata(tab, categorias, metricas, referencia):
    # Get the item name based on the selected tab
    item_name = tab_titles.get(tab, tab)
    
    # If item_name is None, default to tab
    if item_name is None:
        item_name = tab
        
    category_text = summarize_items(categorias, max_items=5)
    metric_text = summarize_items(metricas, max_items=5)
    printed_at = datetime.now().strftime("%d/%m/%Y %H:%M")

    # Adjust item name for player vs player comparison
    if tab == "plyr_vs_plyr":
        item_name = "Comparativa Jugador vs Jugador"

    # Construct the metadata message
    metadata = (
        f"Descargado: {item_name}\n"
        f"Impresión: {printed_at}\n"
        f"Categorías: {category_text}\n"
        f"Métricas: {metric_text}\n"
        f"Comparar por: {referencia}\n"
    )

    return item_name, metadata, printed_at

def build_plyr_vs_plyr(dff, jugador_1, jugador_2, game_tag=None, period_tag=None, metricas=None):
    # Prepare the metrics to use for the radar chart
    metricas = metricas_radar.copy()
    metricas = [m for m in metricas if m in dff.columns]

    # Filter the DataFrame based on game and period tags
    dff_filtrado = dff.copy()
    if game_tag:
        dff_filtrado = dff_filtrado[dff_filtrado["Game Tags"] == game_tag]
    if period_tag:
        dff_filtrado = dff_filtrado[dff_filtrado["Period Tags"] == period_tag]

    # Validate both players are specified
    if not jugador_1 or not jugador_2:
        return go.Figure()

    # Filter for the selected players
    jugadores = [jugador_1, jugador_2]
    dff_jugadores = dff_filtrado[dff_filtrado["Player Name"].isin(jugadores)]

    if dff_jugadores.empty:
        return go.Figure()

    # Calculate average metrics for the selected players
    radar_data = (
        dff_jugadores.groupby("Player Name")[metricas]
        .mean()
        .reset_index()
    )

    # Normalize the radar data
    radar_data_norm = radar_data.copy()

    for m in metricas:
            col_min = radar_data[m].min()
            col_max = radar_data[m].max()
            
            # Normalize only if max > min
            if col_max > col_min:
                radar_data_norm[m] = (radar_data[m] - col_min) / (col_max - col_min)

        # Create the radar chart
    fig = go.Figure()
    colores = ["#48f788", "#89bcef"]

    for idx, row in radar_data_norm.iterrows():
            rgb = tuple(int(colores[idx % len(colores)][1 + i:3 + i], 16) for i in (0, 2, 4))
            
            fig.add_trace(go.Scatterpolar(
                r=row[metricas].values.flatten().tolist(),
                theta=metricas,
                fill="toself",
                name=row["Player Name"],
                mode="markers+lines",
                marker=dict(size=6, color=colores[idx % len(colores)]),
                line=dict(color=colores[idx % len(colores)], width=2),
                fillcolor=f"rgba({rgb[0]},{rgb[1]},{rgb[2]},0.25)",
                text=[f"{val:.2f}" for val in row[metricas].values.flatten().tolist()],
                textposition="top center",
                textfont=dict(size=10, color="#edf1f2")
            ))

        # Update the layout for the radar chart
            fig.update_layout(
            polar=dict(
                radialaxis=dict(
                    visible=True,
                    showline=True,
                    linewidth=1,
                    gridcolor="rgba(200,200,200,0.25)",
                    gridwidth=0.8,
                    tickfont=dict(size=12, color="#edf1f2")
                ),
                angularaxis=dict(
                    tickfont=dict(size=12, color="#edf1f2")
                )
            ),
            showlegend=True,
            template="plotly_dark",
            title=dict(
                text=f"{jugador_1}  ||  {jugador_2}",
                font=dict(size=20, color="#a3e3d0", family="Manrope Light"),
                x=0.5
            ),
            legend=dict(
                font=dict(size=13, color="#edf1f2"),
                orientation="h",
                yanchor="bottom",
                y=-0.25,
                xanchor="center",
                x=0.5,
                bgcolor="rgba(0,0,0,0.4)",
                bordercolor="rgba(137,188,239,0.25)",
                borderwidth=1
            )
        )

    return fig

def build_comparativas(dff, categorias, metricas, referencia):
    # Handle default metrics if none are provided
    metricas = metricas or ["Distance"]
    # Filter metrics to ensure they exist in the DataFrame
    metricas = [m for m in metricas if m in dff.columns]

    # Check if referencia is valid and there are provided metrics
    if referencia not in dff.columns or not metricas:
        return go.Figure()

    # Calculate the average metrics grouped by the reference column
    promedio = (
        dff.groupby(referencia)[metricas]
        .mean()
        .reset_index()
    )

    # Melt the DataFrame to format it for Plotly Express
    promedio_melt = pd.melt(
        promedio,
        id_vars=[referencia],
        value_vars=metricas,
        var_name="Métrica",
        value_name="Valor"
    )

    # Create a horizontal bar chart
    fig = px.bar(
        promedio_melt,
        x="Valor",
        y=referencia,
        color="Métrica",
        orientation="h",
        barmode="group",
        template="plotly_dark",
        color_discrete_sequence=[
            "#edf1f2", "#f1a3fd", "#a3e3d0", "#89bcef", "#48f788", "#f96e83"
        ]
    )

    # Update layout for the chart
    fig.update_layout(
        title={
            "text": f"Comparativo de métricas por {referencia}",
            "font": {
                "color": "#f5f5f5",
                "family": "'Clash Display Semibold', 'Helvetica Neue'",
                "size": 22
            }
        },
        paper_bgcolor="#0b0c0e",
        plot_bgcolor="#0b0c0e",
        font={"color": "#f5f5f5"},
        legend=dict(
            bgcolor="rgba(11,12,14,0.75)",
            bordercolor="#89bcef",
            borderwidth=1
        )
    )

    # Update the x-axis properties
    fig.update_xaxes(
        showgrid=True,
        gridcolor="rgba(137,188,239,0.18)",
        zerolinecolor="rgba(255,255,255,0.08)",
        linecolor="#89bcef",
        tickfont_color="#f5f5f5",
        title_font_color="#a3e3d0"
    )

    # Update the y-axis properties
    fig.update_yaxes(
        showgrid=True,
        gridcolor="rgba(137,188,239,0.18)",
        zerolinecolor="rgba(255,255,255,0.08)",
        linecolor="#89bcef",
        tickfont_color="#f5f5f5",
        title_font_color="#a3e3d0"
    )

    return fig

def build_cronologico(dff, categorias, metricas, referencia):
    metricas = metricas or ["Distance"]
    metricas = [m for m in metricas if m in dff.columns]
    if "Date" not in dff.columns or not metricas:
        return go.Figure()

    dff = dff.dropna(subset=["Date"]).copy()
    if dff.empty:
        return go.Figure()

    cronologico = pd.melt(
        dff,
        id_vars=["Date", "Category"],
        value_vars=metricas,
        var_name="Métrica",
        value_name="Valor"
    )

    fig = px.scatter(
        cronologico,
        x="Date",
        y="Valor",
        color="Category",
        symbol="Métrica",
        color_discrete_sequence=["#edf1f2", "#f1a3fd", "#a3e3d0", "#89bcef", "#48f788", "#f96e83"],
        template="plotly_dark"
    )

    fig.update_traces(
        marker=dict(size=10, line=dict(width=1, color="#ffffff")),
        selector=dict(mode="markers"),
        hoverlabel=dict(bgcolor="#011c24", font_size=12, font_color="#f5f5f5")
    )

    fig.update_layout(
        title={
            "text": f"Evolución cronológica de {' / '.join(metricas)}",
            "font": {"color": "#f5f5f5", "family": "'Clash Display Semibold', 'Helvetica Neue'", "size": 22}
        },
        paper_bgcolor="#0b0c0e",
        plot_bgcolor="#0b0c0e",
        font={"color": "#f5f5f5"},
        legend=dict(bgcolor="rgba(11,12,14,0.75)", bordercolor="#89bcef", borderwidth=1)
    )

    fig.update_xaxes(
        tickformat="%d/%m/%Y",
        showgrid=True,
        gridcolor="rgba(137,188,239,0.18)",
        zerolinecolor="rgba(255,255,255,0.08)",
        linecolor="#89bcef",
        tickfont_color="#f5f5f5",
        title_font_color="#a3e3d0"
    )

    fig.update_yaxes(
        showgrid=True,
        gridcolor="rgba(137,188,239,0.18)",
        zerolinecolor="rgba(255,255,255,0.08)",
        linecolor="#89bcef",
        tickfont_color="#f5f5f5",
        title_font_color="#a3e3d0"
    )

    if LOGO_BASE64:
        fig.add_layout_image(
            dict(
                source="data:image/png;base64," + LOGO_BASE64,
                xref="paper",
                yref="paper",
                x=0.99,
                y=0.01,
                xanchor="right",
                yanchor="bottom",
                sizex=0.12,
                sizey=0.10,
                opacity=0.7,
                layer="above"
            )
        )

    return fig


def truncate_to_n_words(text, n=500):
    words = text.split()
    return " ".join(words[:n])


def build_auto_report_title(categorias, fecha_actividad):
    fecha_text = (
        pd.to_datetime(fecha_actividad).strftime("%d/%m/%Y")
        if fecha_actividad else datetime.now().strftime("%d/%m/%Y")
    )
    if categorias:
        categorias_text = ", ".join(categorias[:3])
        if len(categorias) > 3:
            categorias_text += ", ..."
    else:
        categorias_text = "Todas las categorías"
    return f"Informe {fecha_text} - {categorias_text}"


def section_title(section_value):
    titles = {
        "actividad": "Actividad",
        "actividad_comparativa": "Actividad Comparativa Individual",
        "actividad_promedios": "Actividad/Promedios",
        "acwr": "ACWR",
        "plyr_vs_plyr": "PLYR vs PLYR",
        "comparativas": "Comparativo",
        "cronologico": "Cronológico"
    }
    return titles.get(section_value, section_value)


def build_actividad_report_fig(dff, fecha_dt):
    if "Date" not in dff.columns:
        return go.Figure()
    dff_fecha = dff[dff["Date"].dt.normalize() == fecha_dt]
    metrics = [m for m in ["Distance", "Player Load", "Sprint Distance", "High Speed Distance", "Sprint Efforts"] if m in dff_fecha.columns]
    if dff_fecha.empty or not metrics:
        return go.Figure()

    resumen = dff_fecha.groupby("Player Name")[metrics].sum().reset_index()
    resumen = resumen.sort_values(metrics[0], ascending=False).head(6)
    data = pd.melt(resumen, id_vars=["Player Name"], value_vars=metrics, var_name="Métrica", value_name="Valor")

    fig = px.bar(
        data,
        x="Valor",
        y="Player Name",
        color="Métrica",
        orientation="h",
        barmode="group",
        template="plotly_dark",
        color_discrete_sequence=["#edf1f2", "#f1a3fd", "#a3e3d0", "#89bcef", "#48f788", "#f96e83"]
    )
    fig.update_layout(
        title={"text": f"Actividad {fecha_dt.strftime('%d/%m/%Y')}", "font": {"color": "#f5f5f5", "size": 18}},
        paper_bgcolor="#0b0c0e",
        plot_bgcolor="#0b0c0e",
        font={"color": "#f5f5f5"},
        legend=dict(bgcolor="rgba(11,12,14,0.75)", bordercolor="#89bcef", borderwidth=1)
    )
    return fig


def build_actividad_comparativa_report_fig(dff, fecha_dt):
    if "Date" not in dff.columns:
        return go.Figure()
    metrics = [m for m in ["Distance", "Player Load", "Sprint Distance"] if m in dff.columns]
    dff_fecha = dff[dff["Date"].dt.normalize() == fecha_dt]
    if dff_fecha.empty or not metrics:
        return go.Figure()

    resumen = dff_fecha.groupby("Player Name")[metrics].sum().reset_index()
    dff_acum = dff[dff["Date"].dt.normalize() <= fecha_dt]
    promedio = dff_acum.groupby("Player Name")[metrics].mean().reset_index()
    promedio = promedio.rename(columns={m: f"{m} Prom" for m in metrics})
    tabla = resumen.merge(promedio, on="Player Name", how="left").fillna(0)

    rows = []
    for _, row in tabla.iterrows():
        for m in metrics:
            rows.append({"Player Name": row["Player Name"], "Métrica": m, "Tipo": "Actual", "Valor": row[m]})
            rows.append({"Player Name": row["Player Name"], "Métrica": m, "Tipo": "Promedio", "Valor": row[f"{m} Prom"]})
    if not rows:
        return go.Figure()

    data = pd.DataFrame(rows)
    fig = px.bar(
        data,
        x="Valor",
        y="Player Name",
        color="Tipo",
        facet_col="Métrica",
        orientation="h",
        barmode="group",
        template="plotly_dark",
        category_orders={"Tipo": ["Actual", "Promedio"]},
        color_discrete_sequence=["#48f788", "#89bcef"]
    )
    fig.update_layout(
        title={"text": f"Actividad Comparativa Individual {fecha_dt.strftime('%d/%m/%Y')}", "font": {"color": "#f5f5f5", "size": 18}},
        paper_bgcolor="#0b0c0e",
        plot_bgcolor="#0b0c0e",
        font={"color": "#f5f5f5"}
    )
    fig.for_each_annotation(lambda a: a.update(text=a.text.split('=')[-1]))
    return fig


def build_actividad_promedios_report_fig(dff, fecha_dt):
    metrics = [m for m in metricas_promedios if m in dff.columns]
    dff_fecha = dff[dff["Date"].dt.normalize() == fecha_dt]
    if dff_fecha.empty or not metrics:
        return go.Figure()

    promedio = dff_fecha[metrics].mean().reset_index()
    promedio.columns = ["Métrica", "Valor"]
    fig = px.bar(
        promedio,
        x="Valor",
        y="Métrica",
        orientation="h",
        template="plotly_dark",
        color_discrete_sequence=["#edf1f2"]
    )
    fig.update_layout(
        title={"text": f"Actividad / Promedios {fecha_dt.strftime('%d/%m/%Y')}", "font": {"color": "#f5f5f5", "size": 18}},
        paper_bgcolor="#0b0c0e",
        plot_bgcolor="#0b0c0e",
        font={"color": "#f5f5f5"}
    )
    return fig


def build_acwr_report_fig(dff):
    metrics = [m for m in ["Distance", "Player Load", "Sprint Distance", "High Speed Distance", "Sprint Efforts", "High Speed Efforts", "Impacts"] if m in dff.columns]
    if dff.empty or not metrics:
        return go.Figure()

    ultimos21 = dff["Date"].max() - pd.Timedelta(days=21)
    ultimos7 = dff["Date"].max() - pd.Timedelta(days=7)
    df21 = dff[dff["Date"] >= ultimos21]
    df7 = dff[dff["Date"] >= ultimos7]
    cronica = df21.groupby("Player Name")[metrics].mean().reset_index()
    aguda = df7.groupby("Player Name")[metrics].mean().reset_index()
    tabla = cronica.merge(aguda, on="Player Name", how="outer", suffixes=("_21", "_7")).fillna(0)
    rows = []
    for _, row in tabla.iterrows():
        for m in metrics:
            rows.append({"Player Name": row["Player Name"], "Métrica": m, "Valor": round(row[f"{m}_7"] / row[f"{m}_21"] if row[f"{m}_21"] else 0, 2)})
    data = pd.DataFrame(rows)
    fig = px.bar(
        data,
        x="Valor",
        y="Player Name",
        color="Métrica",
        orientation="h",
        template="plotly_dark",
        color_discrete_sequence=["#edf1f2", "#f1a3fd", "#a3e3d0", "#89bcef", "#48f788", "#f96e83", "#f4c95d"]
    )
    fig.update_layout(
        title={"text": "ACWR - Últimos 7 días vs 21 días", "font": {"color": "#f5f5f5", "size": 18}},
        paper_bgcolor="#0b0c0e",
        plot_bgcolor="#0b0c0e",
        font={"color": "#f5f5f5"}
    )
    return fig


def build_plyr_vs_plyr_report_fig(dff):
    nombres = dff.groupby("Player Name")["Distance"].sum().sort_values(ascending=False).head(2).index.tolist()
    if len(nombres) < 2:
        return go.Figure()
    return build_plyr_vs_plyr(dff, nombres[0], nombres[1], None, None)


def build_section_report_fig(section, dff, fecha_dt, categorias):
    
    if section == "actividad":
        return build_actividad_report_fig(dff, fecha_dt)
    if section == "actividad_comparativa":
        return build_actividad_comparativa_report_fig(dff, fecha_dt)
    if section == "actividad_promedios":
        return build_actividad_promedios_report_fig(dff, fecha_dt)
    if section == "acwr":
        return build_acwr_report_fig(dff)
    if section == "plyr_vs_plyr":
        return build_plyr_vs_plyr_report_fig(dff)
    if section == "comparativas":
        return build_comparativas(dff, categorias, ["Distance"], "Category")
    if section == "cronologico":
        return build_cronologico(dff, categorias, ["Distance"], "Category")
    return go.Figure()


def build_plotly_table(header, rows, title):
    if not rows:
        return None

    columns = [[row.get(col, "") for row in rows] for col in header]
    fig = go.Figure(data=[
        go.Table(
            header=dict(
                values=header,
                fill_color="#1f2c56",
                font=dict(color="white", size=11),
                align="center"
            ),
            cells=dict(
                values=columns,
                fill_color="#0b0c0e",
                font=dict(color="white", size=10),
                align="center"
            )
        )
    ])
    fig.update_layout(
        title={"text": title, "font": {"color": "#f5f5f5", "size": 16}, "x": 0.01},
        width=1600,
        height=900,
        margin=dict(l=10, r=10, t=40, b=10),
        paper_bgcolor="#0b0c0e"
    )
    return fig


def build_section_report_table_fig(section, dff, fecha_dt, categorias):
    if section == "actividad":
        if "Date" not in dff.columns:
            return None
        dff_fecha = dff[dff["Date"].dt.normalize() == fecha_dt]
        metrics = [m for m in ["Distance", "Player Load", "Sprint Distance", "High Speed Distance", "Sprint Efforts"] if m in dff_fecha.columns]
        if dff_fecha.empty or not metrics:
            return None
        resumen = dff_fecha.groupby("Player Name")[metrics].sum().reset_index()
        resumen = resumen.sort_values(metrics[0], ascending=False).head(10).round(2)
        header = ["Player Name"] + metrics
        rows = resumen.to_dict(orient="records")
        return build_plotly_table(header, rows, f"Tabla de Actividad {fecha_dt.strftime('%d/%m/%Y')}")

    if section == "actividad_comparativa":
        if "Date" not in dff.columns:
            return None
        metrics = [m for m in ["Distance", "Player Load", "Sprint Distance"] if m in dff.columns]
        dff_fecha = dff[dff["Date"].dt.normalize() == fecha_dt]
        if dff_fecha.empty or not metrics:
            return None
        resumen = dff_fecha.groupby("Player Name")[metrics].sum().reset_index()
        promedio = dff[dff["Date"].dt.normalize() <= fecha_dt].groupby("Player Name")[metrics].mean().reset_index()
        resumen = resumen.rename(columns={m: f"{m} Actual" for m in metrics})
        promedio = promedio.rename(columns={m: f"{m} Prom" for m in metrics})
        tabla = resumen.merge(promedio, on="Player Name", how="left").fillna(0).round(2)
        header = ["Player Name"] + [f"{m} Actual" for m in metrics] + [f"{m} Prom" for m in metrics]
        rows = tabla.sort_values(f"{metrics[0]} Actual", ascending=False).head(10).to_dict(orient="records")
        return build_plotly_table(header, rows, f"Tabla de Actividad Comparativa Individual {fecha_dt.strftime('%d/%m/%Y')}")

    if section == "actividad_promedios":
        if "Date" not in dff.columns:
            return None
        metrics = [m for m in metricas_promedios if m in dff.columns]
        dff_fecha = dff[dff["Date"].dt.normalize() == fecha_dt]
        if dff_fecha.empty or not metrics:
            return None
        promedio = dff_fecha[metrics].mean().reset_index()
        promedio.columns = ["Métrica", "Valor"]
        promedio["Valor"] = promedio["Valor"].round(2)
        rows = promedio.to_dict(orient="records")
        return build_plotly_table(["Métrica", "Valor"], rows, f"Tabla de Actividad / Promedios {fecha_dt.strftime('%d/%m/%Y')}")

    if section == "acwr":
        metrics = [m for m in ["Distance", "Player Load", "Sprint Distance", "High Speed Distance", "Sprint Efforts", "High Speed Efforts", "Impacts"] if m in dff.columns]
        if dff.empty or not metrics:
            return None
        ultimos21 = dff["Date"].max() - pd.Timedelta(days=21)
        ultimos7 = dff["Date"].max() - pd.Timedelta(days=7)
        df21 = dff[dff["Date"] >= ultimos21]
        df7 = dff[dff["Date"] >= ultimos7]
        cronica = df21.groupby("Player Name")[metrics].mean().reset_index()
        aguda = df7.groupby("Player Name")[metrics].mean().reset_index()
        tabla = cronica.merge(aguda, on="Player Name", how="outer", suffixes=("_21", "_7")).fillna(0)
        rows = []
        for _, row in tabla.iterrows():
            item = {"Player Name": row["Player Name"]}
            for m in metrics:
                item[f"{m} ACWR"] = round((row[f"{m}_7"] / row[f"{m}_21"]) if row[f"{m}_21"] else 0, 2)
            rows.append(item)
        if not rows:
            return None
        header = ["Player Name"] + [f"{m} ACWR" for m in metrics]
        df_ratio = pd.DataFrame(rows).sort_values(f"{metrics[0]} ACWR", ascending=False).head(10)
        return build_plotly_table(header, df_ratio.to_dict(orient="records"), "Tabla ACWR (7 días vs 21 días)")

    return None


def fig_to_png_bytes(fig, width=1600, height=900, scale=2, timeout=10):
    """Convierte una figura Plotly en PNG usando Kaleido."""
    if fig is None or not getattr(fig, "data", None):
        return None

    # 1) Kaleido (preferido)
    try:
        import plotly.io as pio
        png = pio.to_image(fig, format="png", width=width, height=height, scale=scale)
        if png:
            return png
    except Exception as e:
        logging.debug("Kaleido falló: %s", e)

    # 2) WeasyPrint: renderizar HTML de la figura y convertir a PNG vía PDF intermedio
    try:
        if WeasyHTML is not None:
            html_str = fig.to_html(full_html=False, include_plotlyjs="cdn")
            # envolver en un HTML simple para WeasyPrint
            wrapper = f"""
            <html><head><meta charset="utf-8"></head>
            <body style="background:white;margin:0;padding:0;">
            {html_str}
            </body></html>
            """
            # WeasyPrint puede generar PDF; para PNG generamos PDF y luego lo convertimos con Pillow
            pdf_bytes = WeasyHTML(string=wrapper).write_pdf()
            # convertir primera página del PDF a PNG con Pillow (si Pillow soporta)
            try:
                from pdf2image import convert_from_bytes
                pages = convert_from_bytes(pdf_bytes)
                buf = io.BytesIO()
                pages[0].save(buf, format="PNG")
                return buf.getvalue()
            except Exception:
                # si no está pdf2image, devolver el PDF bytes (caller puede usar build_graph_pdf_bytes)
                logging.debug("pdf2image no disponible para convertir PDF->PNG")
    except Exception as e:
        logging.debug("WeasyPrint fallback falló: %s", e)

    # 3) Selenium screenshot (solo si la app es accesible y Selenium está instalado)
    try:
        from selenium.webdriver.common.by import By

        chrome_options = Options()
        chrome_options.add_argument("--headless=new")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=chrome_options)
        try:
            # Debés exponer la app (ej: http://127.0.0.1:8050) y conocer el selector del div del gráfico
            app_url = os.environ.get("APP_URL", "http://127.0.0.1:8050")
            graph_selector = os.environ.get("GRAPH_SELECTOR", "#report_figures_preview .tab-graph")  # ajustar
            driver.set_page_load_timeout(timeout)
            driver.get(app_url)
            time.sleep(1.0)
            el = driver.find_element(By.CSS_SELECTOR, graph_selector)
            png = el.screenshot_as_png
            return png
        finally:
            driver.quit()
    except Exception as e:
        logging.debug("Selenium fallback no disponible o falló: %s", e)

    # Si todo falla
    logging.warning("No se pudo generar PNG para la figura (todos los fallbacks fallaron).")
    return None

def build_graph_pdf_from_fig(fig, width=1600, height=900, scale=2):

    png_bytes = fig_to_png_bytes(fig, width=width, height=height, scale=scale)
    if png_bytes is None:
        return None

    title = getattr(fig, "name", "Grafico") or "Grafico"
    return build_graph_pdf_bytes(title=title, fig_png=png_bytes)


def build_graph_pdf_bytes(title, fig_png, page_size=A4, margin=inch * 0.75):
    """Construye un PDF con una imagen PNG incrustada y devuelve bytes."""
    if fig_png is None:
        return None
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=page_size)
    width, height = page_size

    # Header
    c.setFont("Helvetica-Bold", 16)
    y = height - margin
    c.drawString(margin, y, title)
    y -= 24

    try:
        image = ImageReader(io.BytesIO(fig_png))
        img_width, img_height = image.getSize()
        max_width = width - 2 * margin
        max_height = y - margin
        ratio = min(max_width / img_width, max_height / img_height, 1)
        draw_width = img_width * ratio
        draw_height = img_height * ratio

        if draw_width <= 0 or draw_height <= 0:
            raise ValueError("Imagen inválida para PDF")

        if y - draw_height < margin:
            c.showPage()
            y = height - margin
            c.setFont("Helvetica-Bold", 16)
            c.drawString(margin, y, title)
            y -= 24

        c.drawImage(image, margin, y - draw_height, width=draw_width, height=draw_height, preserveAspectRatio=True, mask='auto')
        y -= draw_height + 12
        c.setFont("Helvetica", 9)
        c.drawString(margin, margin / 2, "Generado con ReportLab")
    except Exception as e:
        logging.warning("No se pudo embebar PNG en PDF: %s", e)
        return None

    c.save()
    buffer.seek(0)
    return buffer.read()


def combine_image_bytes_vertically(image_bytes_list, spacing=20, background=(255, 255, 255, 255)):
    """Combina varias imágenes (bytes PNG) verticalmente y devuelve PNG bytes."""
    if not image_bytes_list:
        return None
    images = []
    for image_bytes in image_bytes_list:
        try:
            img = PILImage.open(io.BytesIO(image_bytes)).convert("RGBA")
            images.append(img)
        except Exception as e:
            logging.warning("No se pudo abrir imagen para combinar: %s", e)
            return None

    max_width = max(img.width for img in images)
    total_height = sum(img.height for img in images) + spacing * (len(images) - 1)
    combined = PILImage.new("RGBA", (max_width, total_height), background)

    y = 0
    for img in images:
        x = (max_width - img.width) // 2
        combined.paste(img, (x, y), img)
        y += img.height + spacing

    buf = io.BytesIO()
    combined.convert("RGB").save(buf, format="PNG")
    return buf.getvalue()

def draw_wrapped_text(c, text, x, y, width, leading, page_height, margin, header_func=None):
    font_name = "Manrope-Light" if "Manrope-Light" in pdfmetrics.getRegisteredFontNames() else "Helvetica"
    text_obj = c.beginText(x, y)
    text_obj.setFont(font_name, 10)
    for paragraph in text.split("\n"):
        lines = textwrap.wrap(paragraph, width=100)
        if not lines:
            lines = [""]
        for line in lines:
            if y < margin + 60:
                c.drawText(text_obj)
                c.showPage()
                if header_func:
                    y = header_func()
                else:
                    y = page_height - margin
                text_obj = c.beginText(x, y)
                text_obj.setFont(font_name, 10)
            text_obj.textLine(line)
            y -= leading
    c.drawText(text_obj)
    return y


def draw_page_header(c, title, author, fecha_text, filters_text, logo_bytes, width, height, margin):
    y = height - margin
    title_font = "ClashDisplay-Semibold" if "ClashDisplay-Semibold" in pdfmetrics.getRegisteredFontNames() else "Helvetica-Bold"
    small_font = "Manrope-Light" if "Manrope-Light" in pdfmetrics.getRegisteredFontNames() else "Helvetica"

    c.setFont(title_font, 22)
    c.drawString(margin, y, title)
    if logo_bytes:
        try:
            logo = load_image_reader_from_bytes(logo_bytes)
            c.drawImage(logo, width - margin - 100, height - margin - 60, width=100, height=60, preserveAspectRatio=True, mask='auto')
        except Exception as e:
            logging.warning("No se pudo dibujar el logo en PDF: %s", e)
    y -= 22
    c.setFont(small_font, 9)
    c.drawString(margin, y, fecha_text)
    if filters_text:
        y = draw_wrapped_text(c, filters_text, margin, y - 12, width - 2 * margin, 11, height, margin)
    y -= 14
    c.setStrokeColorRGB(0.6, 0.7, 0.8)
    c.setLineWidth(0.5)
    c.line(margin, y, width - margin, y)
    return y - 12


def draw_page_footer(c, author, width, margin):
    footer_font = "ClashDisplay-Semibold" if "ClashDisplay-Semibold" in pdfmetrics.getRegisteredFontNames() else "Helvetica-Bold"
    c.setFont(footer_font, 9)
    footer_text = f"Creado por {author}"
    footer_y = margin / 2
    c.drawRightString(width - margin, footer_y, footer_text)


def draw_page_header_and_footer(c, title, author, fecha_text, filters_text, logo_bytes, width, height, margin):
    y = draw_page_header(c, title, author, fecha_text, filters_text, logo_bytes, width, height, margin)
    draw_page_footer(c, author, width, margin)
    return y


def load_image_reader_from_bytes(image_bytes):
    try:
        return ImageReader(io.BytesIO(image_bytes))
    except Exception:
        if PILImage is not None:
            try:
                pil_img = PILImage.open(io.BytesIO(image_bytes))
                return ImageReader(pil_img)
            except Exception:
                pass
        raise


def build_report_pdf_multi(title, author, logo_bytes, sections, fecha_text, filters_text=None, page_size=A4, margin=inch * 0.6):
    """
    Construye un PDF multi-página que incluye:
    - Portada con título, autor, fecha y filtros.
    - Para cada sección: título, texto envuelto y las imágenes (figura y/o tabla) incrustadas.
    - Las imágenes se escalan manteniendo el aspecto; si no caben en la página se crea una nueva.
    Devuelve bytes del PDF final.
    """
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=page_size)
    width, height = page_size

    # Fuentes a usar (registradas previamente si están disponibles)
    title_font = "ClashDisplay-Semibold" if "ClashDisplay-Semibold" in pdfmetrics.getRegisteredFontNames() else "Helvetica-Bold"
    small_font = "Manrope-Light" if "Manrope-Light" in pdfmetrics.getRegisteredFontNames() else "Helvetica"

    def draw_header(c):
        """Dibuja el encabezado y devuelve la coordenada y disponible debajo del encabezado."""
        y = height - margin
        c.setFont(title_font, 18)
        c.setFillColorRGB(0, 0, 0)
        c.drawString(margin, y, title)
        if logo_bytes:
            try:
                logo = load_image_reader_from_bytes(logo_bytes)
                c.drawImage(logo, width - margin - 100, height - margin - 60, width=100, height=60, preserveAspectRatio=True, mask='auto')
            except Exception:
                pass
        y -= 26
        c.setFont(small_font, 9)
        c.drawString(margin, y, f"Autor: {author}")
        c.drawRightString(width - margin, y, f"Fecha: {fecha_text}")
        y -= 14
        if filters_text:
            # draw_wrapped_text devuelve la nueva y después de escribir el texto
            y = draw_wrapped_text(c, filters_text, margin, y, int(width - 2 * margin), 11, height, margin, header_func=None)
        y -= 8
        c.setStrokeColorRGB(0.6, 0.7, 0.8)
        c.setLineWidth(0.5)
        c.line(margin, y, width - margin, y)
        return y - 12

    def draw_footer(c, page_num):
        """Dibuja el pie de página con número de página."""
        footer_font = title_font if title_font else "Helvetica-Bold"
        c.setFont(footer_font, 9)
        footer_text = f"Creado por {author} — Página {page_num}"
        c.drawRightString(width - margin, margin / 2, footer_text)

    # --- Portada ---
    c.setFont(title_font, 28)
    c.drawCentredString(width / 2, height - 2.5 * inch, title)
    c.setFont(small_font, 12)
    c.drawCentredString(width / 2, height - 2.5 * inch - 28, f"Autor: {author} | Fecha: {fecha_text}")
    if filters_text:
        c.setFont(small_font, 10)
        draw_wrapped_text(c, filters_text, margin, height - 2.5 * inch - 60, int(width - 2 * margin), 11, height, margin)
    c.showPage()

    page_num = 1

    # --- Iterar secciones ---
    for section in sections:
        page_num += 1
        y = draw_header(c)

        # Título de sección
        c.setFont(title_font, 16)
        c.drawString(margin, y, section.get("title", ""))
        y -= 20

        # Texto de la sección (envuelto)
        c.setFont(small_font, 10)
        text = section.get("text", "")
        if text:
            y = draw_wrapped_text(
                c,
                text,
                margin,
                y,
                int(width - 2 * margin),
                12,
                height,
                margin,
                header_func=None
            )
            y -= 8

        # Insertar imágenes: primero figura principal, luego tabla si existe
        for img_key, caption_key, max_h_default in [
            ("img", "caption", 900),
            ("table_img", "table_caption", 900)
        ]:
            img_bytes = section.get(img_key)
            caption = section.get(caption_key, "")
            if not img_bytes:
                continue
            try:
                image = load_image_reader_from_bytes(img_bytes)
                img_w, img_h = image.getSize()
                max_w = width - 2 * margin
                max_h = y - margin
                # Si no hay suficiente espacio vertical, abrir nueva página y redibujar encabezado
                if max_h <= 120:
                    c.showPage()
                    page_num += 1
                    y = draw_header(c)
                    c.setFont(title_font, 16)
                    c.drawString(margin, y, section.get("title", ""))
                    y -= 20
                    max_h = y - margin

                # limitar altura a un valor razonable para evitar imágenes gigantes
                max_h = min(max_h, max_h_default)
                ratio = min(max_w / img_w, max_h / img_h, 1)
                draw_w = img_w * ratio
                draw_h = img_h * ratio
                x = margin + (max_w - draw_w) / 2
                c.drawImage(image, x, y - draw_h, width=draw_w, height=draw_h, preserveAspectRatio=True, mask='auto')
                y -= draw_h + 6

                if caption:
                    c.setFont(small_font, 9)
                    c.drawString(margin, y, caption)
                    y -= 14

            except Exception as e:
                logging.warning("No se pudo incrustar imagen en sección '%s': %s", section.get("title", ""), e)
                # dejar un pequeño espacio y continuar
                y -= 8

        # Después de insertar contenido, si queda poco espacio, crear nueva página
        if y < margin + 120:
            c.showPage()
            page_num += 1

    # Pie de página en la última página
    draw_footer(c, page_num)
    c.save()
    buffer.seek(0)
    return buffer.read()

def build_graph_html_pdf(title, fig_png, table_png=None, filters_text=None, logo_bytes=None):
    if WeasyHTML is None:
        raise ImportError("WeasyPrint no está instalado. Instalar weasyprint para generar PDF desde HTML.")

    logo_html = ""
    if logo_bytes is not None:
        logo_base64 = base64.b64encode(logo_bytes).decode("utf-8")
        logo_html = f'<img class="report-logo" src="data:image/png;base64,{logo_base64}" alt="Logo" />'

    fig_b64 = base64.b64encode(fig_png).decode("utf-8")
    table_html = ""
    if table_png is not None:
        table_b64 = base64.b64encode(table_png).decode("utf-8")
        table_html = f"<div class='report-table'><img class='report-image' src='data:image/png;base64,{table_b64}' alt='Tabla' /></div>"

    filters_html = html_module.escape(filters_text or "").replace("\n", "<br />")

    html_content = f"""
    <html>
    <head>
      <meta charset='utf-8' />
      <style>
        body {{ font-family: Arial, sans-serif; margin: 0; padding: 0; background: #fff; color: #222; }}
        .container {{ width: 100%; max-width: 1100px; margin: 0 auto; padding: 24px; }}
        .header {{ text-align: center; margin-bottom: 24px; }}
        .report-logo {{ max-height: 70px; margin-bottom: 16px; }}
        h1 {{ font-size: 28px; color: #1c3d72; margin: 0 0 8px; }}
        .meta {{ font-size: 12px; color: #555; margin-bottom: 18px; }}
        .report-section {{ margin-bottom: 30px; page-break-inside: avoid; }}
        .report-section h2 {{ font-size: 20px; margin-bottom: 10px; color: #1f4a7d; }}
        .report-image, .report-table-image {{ width: 100%; border: 1px solid #ccc; border-radius: 10px; margin-bottom: 10px; }}
        .caption {{ font-size: 12px; color: #555; margin: 0 0 12px; }}
        .report-text {{ font-size: 13px; line-height: 1.6; color: #333; }}
      </style>
    </head>
    <body>
      <div class='container'>
        <div class='header'>
          {logo_html}
          <h1>{html_module.escape(title)}</h1>
          <div class='meta'>{filters_html}</div>
        </div>
        <div class='report-section'>
          <h2>{html_module.escape(title)}</h2>
          <div class='report-image-container'>
            <img class='report-image' src='data:image/png;base64,{fig_b64}' alt='Figura' />
          </div>
          {table_html}
        </div>
      </div>
    </body>
    </html>
    """

    return WeasyHTML(string=html_content).write_pdf()


import tempfile
from pathlib import Path
from datetime import datetime, timedelta

def save_pdf_bytes_to_temp_file(pdf_bytes, filename=None):
    """Save PDF bytes to a temporary file and return the file path."""
    
    # Default filename if none is provided
    if filename is None:
        filename = "reporte_temporal.pdf"
    
    # Create a path for the temporary file
    temp_path = Path(tempfile.gettempdir()) / filename
    
    # Write the bytes to the temporary file
    temp_path.write_bytes(pdf_bytes)
    
    return temp_path

# Get the last update timestamp, formatted for display
ultima_actualizacion = (
    datetime.now() - timedelta(hours=3)
).strftime("%d/%m/%Y - %H:%M")

app.layout = html.Div([

    # ===== TÍTULO =====

    html.Div(
        [
            html.Div(
                [
                    html.Div(
                        [
                            html.Img(
                                src="/assets/logo_dataload_2.png",
                                style={
                                    "width": "90%",
                                    "height": "auto",
                                    "objectFit": "contain",
                                    "marginBottom": "10px",
                                    "border": "1px solid #011c24",
                                    "borderRadius": "12px"
                                }
                            )
                        ],
                        style={
                            "flex": "0 0 160px",
                            "display": "flex",
                            "alignItems": "center",
                            "justifyContent": "flex-start"
                        }
                    ),
                    html.Div(
                        [
                            html.H1(
                                "CARGA EXTERNA - DANUBIO FORMATIVAS 2026",
                                style={
                                    "color": "#ffffff",
                                    "textAlign": "center",
                                    "fontSize": "36px",
                                    "fontWeight": "700",
                                    "fontFamily": "'Clash Display Semibold'",
                                    "lineHeight": "1.50",
                                    "letterSpacing": "0.02em",
                                    "margin": "0",
                                    "linecolor": "#89bcef",
                                    "boxSizing": "border-box"
                                }
                            ),
                            html.P(
                                "DATA LOAD - Plataforma de Análisis Deportivo",
                                style={
                                    "color": "#d0f0d9",
                                    "margin": "10px 0 0",
                                    "fontSize": "15px",
                                    "textAlign": "center",
                                    "fontFamily": "'Manrope Light', sans-serif",
                                    "lineHeight": "1.4"
                                }
                            )
                        ],
                        style={
                            "flex": "1",
                            "minWidth": "2px"
                        }
                    )
                ],
                style={
                    "flex":"1",
                    "height":"90px",
                    "width": "80%",
                    "padding": "18px",
                    "display": "flex",
                    "alignItems": "center",
                    "justifyContent": "center",
                    "gap": "18px",
                    "boxSizing": "border-box"
                }
            )
        ],
        style={
            "width": "100%",
            "paddingTop": "10px",
            "paddingBottom": "12px"
        }
    ),

    # ===== CONTENIDO GENERAL =====

    html.Div([

        # ==================================================
        # SIDEBAR IZQUIERDO
        # ==================================================

        html.Div([

            # MENÚ

            html.Div([

                html.H2(
                    "MENÚ",
                    style={
                        "color":"#edf1f2",
                        "fontSize":"22px",
                        "fontWeight":"700",
                        "textAlign":"center",
                        "marginBottom":"20px"
                    }
                ),

                dcc.Tabs(
    id="tabs",
    value="actividad_promedios",
    vertical=True,
    className="tab-menu",
    style={
        "width":"100%",
        "background":"transparent",
        "border":"none"
    },

    children=[

        dcc.Tab( 
            label="ACTIVIDAD",
            value="actividad",
            className="tab-item",
            selected_className="tab-item-selected",
            style={
                "whiteSpace": "normal",
                "overflow": "hidden",
                "textOverflow": "ellipsis",
                "color":"#edf1f2",
                "fontSize":"11px",
                "textAlign":"center",
                "fontWeight":"600",
                "borderTop":"1px solid rgba(137,188,239,.18)",
                "padding":"2px 6px",       
                "marginBottom":"1px"       
            },
            selected_style={
                "whiteSpace": "normal",
                "overflow": "hidden",
                "textOverflow": "ellipsis",
                "color":"#a3e3d0",
                "fontSize":"11px",
                "textAlign":"center",
                "fontWeight":"600",
                "borderTop":"1px solid #a3e3d0",
                "padding":"2px 6px",
                "backgroundColor":"#011c24",
                "marginBottom":"1px"
            }
        ),

        dcc.Tab(
            label="ACTIVIDAD COMPARATIVA INDIVIDUAL",
            value="actividad_comparativa",
            className="tab-item",
            selected_className="tab-item-selected",
            style={
                "whiteSpace": "normal",
                "overflow": "hidden",
                "textOverflow": "ellipsis",
                "color":"#edf1f2",
                "fontSize":"11px",
                "textAlign":"center",
                "fontWeight":"600",
                "borderTop":"1px solid rgba(137,188,239,.18)",
                "padding":"2px 6px",       
                "marginBottom":"1px"       
            },
            selected_style={
                "whiteSpace": "normal",
                "overflow": "hidden",
                "textOverflow": "ellipsis",
                "color":"#a3e3d0",
                "fontSize":"11px",
                "textAlign":"center",
                "fontWeight":"600",
                "borderTop":"1px solid #a3e3d0",
                "padding":"2px 6px",
                "backgroundColor":"#011c24",
                "marginBottom":"1px"
            }
        ),

        dcc.Tab(
            label="ACTIVIDAD/PROMEDIOS",
            value="actividad_promedios",
            className="tab-item",
            selected_className="tab-item-selected",
            style={
                "whiteSpace": "normal",
                "overflow": "hidden",
                "textOverflow": "ellipsis",
                "color":"#edf1f2",
                "fontSize":"11px",
                "textAlign":"center",
                "fontWeight":"600",
                "borderTop":"1px solid rgba(137,188,239,.18)",
                "padding":"2px 6px",       
                "marginBottom":"1px"       
            },
            selected_style={
                "whiteSpace": "normal",
                "overflow": "hidden",
                "textOverflow": "ellipsis",
                "color":"#a3e3d0",
                "fontSize":"11px",
                "textAlign":"center",
                "fontWeight":"600",
                "borderTop":"1px solid #a3e3d0",
                "padding":"2px 6px",
                "backgroundColor":"#011c24",
                "marginBottom":"1px"
            }
        ),

        dcc.Tab(
            label="ACWR",
            value="acwr",
            className="tab-item",
            selected_className="tab-item-selected",
            style={
                "whiteSpace": "normal",
                "overflow": "hidden",
                "textOverflow": "ellipsis",
                "color":"#edf1f2",
                "fontSize":"11px",
                "textAlign":"center",
                "fontWeight":"600",
                "borderTop":"1px solid rgba(137,188,239,.18)",
                "padding":"2px 6px",       
                "marginBottom":"1px"       
            },
            selected_style={
                "whiteSpace": "normal",
                "overflow": "hidden",
                "textOverflow": "ellipsis",
                "color":"#a3e3d0",
                "fontSize":"11px",
                "textAlign":"center",
                "fontWeight":"600",
                "borderTop":"1px solid #a3e3d0",
                "padding":"2px 6px",
                "backgroundColor":"#011c24",
                "marginBottom":"1px"
            }
        ),
        
        dcc.Tab(
            label="PLYR vs PLYR",
            value="plyr_vs_plyr",
            className="tab-item",
            selected_className="tab-item-selected",
            style={
                "whiteSpace": "normal",
                "overflow": "hidden",
                "textOverflow": "ellipsis",
                "color":"#edf1f2","fontSize":"12px","textAlign":"center","fontWeight":"600",
                "borderTop":"1px solid rgba(137,188,239,.18)",
                "padding":"2px 6px","marginBottom":"1px"
            },
            selected_style={
                "whiteSpace": "normal",
                "overflow": "hidden",
                "textOverflow": "ellipsis",
                "color":"#a3e3d0","fontSize":"12px","textAlign":"center","fontWeight":"600",
                "borderTop":"1px solid #a3e3d0",
                "padding":"2px 6px","backgroundColor":"#011c24","marginBottom":"1px"
            }
        ),

        dcc.Tab(
            label="COMPARATIVO",
            value="comparativas",
            className="tab-item",
            selected_className="tab-item-selected",
            style={
                "whiteSpace": "normal",
                "overflow": "hidden",
                "textOverflow": "ellipsis",      
                "color":"#edf1f2",
                "fontSize":"11px",
                "textAlign":"center",
                "fontWeight":"600",
                "borderTop":"1px solid rgba(137,188,239,.18)",
                "padding":"2px 6px",       
                "marginBottom":"1px"       
            },
            selected_style={
                "whiteSpace": "normal",
                "overflow": "hidden",
                "textOverflow": "ellipsis",
                "color":"#a3e3d0",
                "fontSize":"11px",
                "textAlign":"center",
                "fontWeight":"600",
                "borderTop":"1px solid #a3e3d0",
                "padding":"2px 6px",
                "backgroundColor":"#011c24",
                "marginBottom":"1px"
            }
        ),

        dcc.Tab(
            label="CRONOLÓGICO",
            value="cronologico",
            className="tab-item",
            selected_className="tab-item-selected",
            style={
                "whiteSpace": "normal",
                "overflow": "hidden",
                "textOverflow": "ellipsis",      
                "color":"#edf1f2",
                "fontSize":"11px",
                "textAlign":"center",
                "fontWeight":"600",
                "borderTop":"1px solid rgba(137,188,239,.18)",
                "padding":"2px 6px",       
                "marginBottom":"1px"       
            },
            selected_style={
                "whiteSpace": "normal",
                "overflow": "hidden",
                "textOverflow": "ellipsis",
                "color":"#a3e3d0",
                "fontSize":"11px",
                "textAlign":"center",
                "fontWeight":"600",
                "borderTop":"1px solid #a3e3d0",
                "padding":"2px 6px",
                "backgroundColor":"#011c24",
                "marginBottom":"1px"
            }
        ),
        dcc.Tab(
            label="INFORME",
            value="informe",
            className="tab-item",
            selected_className="tab-item-selected",
            style={
                "whiteSpace": "normal",
                "overflow": "hidden",
                "textOverflow": "ellipsis",
                "color":"#edf1f2",
                "fontSize":"11px",
                "textAlign":"center",
                "fontWeight":"600",
                "borderTop":"1px solid rgba(137,188,239,.18)",
                "padding":"2px 6px",       
                "marginBottom":"1px"       
            },
            selected_style={
                "whiteSpace": "normal",
                "overflow": "hidden",
                "textOverflow": "ellipsis",
                "color":"#a3e3d0",
                "fontSize":"11px",
                "textAlign":"center",
                "fontWeight":"600",
                "borderTop":"1px solid #a3e3d0",
                "padding":"2px 6px",
                "backgroundColor":"#011c24",
                "marginBottom":"1px"
            }
        )

    ]
),

                html.Div(
                    f"Última actualización: {ultima_actualizacion}",
                    style={
                        "color":"#dcdcdc",
                        "fontSize":"9px",
                        "marginTop":"12px",
                        "padding":"0px 3px",
                        "textAlign":"center"
                    }
                )

            ],

            style={
                "padding":"2px",
                "width":"160px",
                "background":"linear-gradient(180deg,#000c0f,#011c24)",
                "borderRadius":"12px",
                "border":"1px solid rgba(137,188,239,.18)"
                })
            ],

        style={
            "display":"flex",
            "flexDirection":"column",
            "width":"120px",
            "minWidth":"160px",
            "minHeight":"45vh",
            "gap":"2px 4px",
            "position":"sticky",
            "top":"100px",
            "alignSelf":"flex-start",
            "alignItems":"center"
        }),

        # ==================================================
        # CONTENIDO CENTRAL
        # ==================================================

        html.Div([

            dcc.Download(id="download-graph"),
            dcc.Download(id="download-table"),
            dcc.Download(id="download-report"),

            # GRÁFICO

            html.Div(    id="contenido-tab",
    className="tab-content",
    style={
        "display":"flex",
        "flex":"1",
        "padding":"24px",
        "gap":"18px",
        "width":"100%",
        "maxWidth":"1000px",
        "minWidth":"1000px",
        "height":"900px",

        "overflowX":"hidden",
        "overflowY":"hidden",

        "position":"relative",
        "boxShadow":"0 18px 40px rgba(1, 28, 36, 1)",
    }
),

            # BOTONES DE DESCARGA

          html.Div([
                 # Descarga de tablas (CSV / XLSX / PNG / PDF)
    html.Button(
        html.Img(src="/assets/icon-download-csv.svg", style={"width":"16px"}),
        id="download-table-csv",
        className="download-btn",
        n_clicks=0
    ),
    html.Button(
        html.Img(src="/assets/icon-download-xlsx.svg", style={"width":"16px"}),
        id="download-table-xlsx",
        className="download-btn",
        n_clicks=0
    ),
    html.Button(
        html.Img(src="/assets/icon-download-png.svg", style={"width":"16px"}),
        id="download-table-png",
        className="download-btn",
        n_clicks=0
    ),
    html.Button(
        html.Img(src="/assets/icon-download-pdf.svg", style={"width":"16px"}),
        id="download-table-pdf",
        className="download-btn",
        n_clicks=0
    ),

    # Descarga del informe completo (opcional)
    html.Button(
        html.Img(src="/assets/icon-download-pdf.svg", style={"width":"16px"}),
        id="generate_report",
        className="download-btn",
        n_clicks=0
    ),
],
style={
    "display":"flex",
    "justifyContent":"center",
    "gap":"12px",
    "marginTop":"10px",
    "paddingBottom":"18px"
})

        ],

        style={
            "display":"flex",
            "flexDirection":"column",
            "flex":"1",
            "gap":"18px"
        }),

            html.Div(
                [
                    html.Div(
                        [
                            html.H4(
                                "Categorías",
                                style={
                                    "color":"#a3e3d0",
                                    "fontSize":"13px",
                                    "fontWeight":"600",
                                    "fontFamily":"'Clash Display Semibold', 'Helvetica Neue'",
                                    "marginBottom":"12px"
                                }
                            ),
                            dcc.Checklist(
                                id="categoria",
                                options=[
                                    {"label":c, "value":c}
                                    for c in sorted(df["Category"].dropna().unique())
                                ],
                                inline=False,
                                style={
                                    "display": "grid",
                                    "gridTemplateColumns": "repeat(2, minmax(0, 1fr))",
                                    "gap": "8px"
                                },
                                labelStyle={
                                    "color":"white",
                                    "fontSize":"10px",
                                    "marginBottom":"6px"
                                }
                            )
                        ],
                        className="filter-card",
                        style={
                            "padding":"14px"
                        }
                    ),
                    html.Div(
                        [
                            html.H4(
                                "Fecha de actividad",
                                style={
                                    "color":"#a3e3d0",
                                    "fontSize":"13px",
                                    "fontWeight":"500",
                                    "fontFamily":"'Clash Display Semibold', 'Helvetica Neue'",
                                    "padding":"12px 12px",
                                    "marginBottom":"12px"
                                }
                            ),
                            dcc.DatePickerSingle(
                                id="fecha-actividad",
                                date=fecha_max.date(),
                                min_date_allowed=df["Date"].min().date(),
                                max_date_allowed=df["Date"].max().date(),
                                display_format="DD/MM/YYYY",
                                style={
                                    "width": "100%",
                                    "backgroundColor": "#011c24",
                                    "color": "#f5f5f5",
                                    "border": "1px solid rgba(137,188,239,0.22)",
                                    "borderRadius": "8px",
                                    "padding": "8px 8px"
                                }
                            )
                        ],
                        id="actividad-fecha-container",
                        className="filter-card",
                        style={
                            "display": "none",
                            "padding":"8px 8px",
                            "marginBottom":"2px"
                        }
                    ),
                    html.Div(
                        [
                            html.H4(
                                "Métricas",
                                style={
                                    "color":"#a3e3d0",
                                    "fontSize":"13px",
                                    "fontWeight":"600",
                                    "fontFamily":"'Clash Display Semibold', 'Helvetica Neue'",
                                    "marginBottom":"12px"
                                }
                            ),
                            dcc.Checklist(
                                id="metrica",
                                options=[
                                    {"label":m, "value":m}
                                    for m in metricas
                                ],
                                value=["Distance"],
                                inline=False,
                                style={
                                    "display": "grid",
                                    "gridTemplateColumns": "repeat(2, minmax(0, 1fr))",
                                    "gap": "5px"
                                },
                                labelStyle={
                                    "color":"white",
                                    "fontSize":"10px",
                                    "marginBottom":"2px"
                                }
                            )
                        ],
                        className="filter-card",
                        style={
                            "padding":"14px"
                        }
                    ),
                    html.Div(
                        [
                            html.H4(
                                "Comparar por",
                                style={
                                    "color":"#a3e3d0",
                                    "fontSize":"13px",
                                    "fontWeight":"600",
                                    "fontFamily":"'Clash Display Semibold', 'Helvetica Neue'",
                                    "marginBottom":"12px"
                                }
                            ),
                            dcc.RadioItems(
                                id="referencia",
                                options=[
                                    {"label":r, "value":r}
                                    for r in referencias
                                ],
                                value="Category",
                                style={
                                    "display": "grid",
                                    "gridTemplateColumns": "repeat(2, minmax(0, 1fr))",
                                    "gap": "10px"
                                },
                                labelStyle={
                                    "color":"white",
                                    "fontSize":"10px",
                                    "marginBottom":"2px"
                                }
                            )
                        ],
                        className="filter-card",
                        style={
                            "padding":"16px"
                        }
                    ),
                    html.Div(
                        [
                            html.H4(
                                "Filtrar por",
                                style={
                                    "color":"#a3e3d0",
                                    "fontSize":"13px",
                                    "fontWeight":"600",
                                    "fontFamily":"'Clash Display Semibold', 'Helvetica Neue'",
                                    "marginBottom":"12px"
                                }
                            ),

                            html.Div(
                                [
                                    html.Div(
                                        [
                                            html.P(
                                                "Jugador",
                                                style={
                                                    "color":"#f5f5f5",
                                                    "fontSize":"10px",
                                                    "marginBottom":"6px"
                                                }
                                            ),
                                            dcc.Dropdown(
                                                id="jugador",
                                                options=[
                                                    {"label": x, "value": x}
                                                    for x in sorted(
                                                        df["Player Name"]
                                                        .dropna()
                                                        .unique()
                                                    )
                                                ],
                                                multi=True
                                            )
                                        ]
                                    ),

                                    html.Div(
                                        [
                                            html.P(
                                                "Athlete Tags",
                                                style={
                                                    "color":"#f5f5f5",
                                                    "fontSize":"10px",
                                                    "marginBottom":"6px"
                                                }
                                            ),
                                            dcc.Dropdown(
                                                id="athlete",
                                                options=[
                                                    {"label": x, "value": x}
                                                    for x in sorted(
                                                        df["Athlete Tags"]
                                                        .dropna()
                                                        .unique()
                                                    )
                                                ],
                                                multi=True
                                            )
                                        ]
                                    ),

                                    html.Div(
                                        [
                                            html.P(
                                                "Game Tags",
                                                style={
                                                    "color":"#f5f5f5",
                                                    "fontSize":"10px",
                                                    "marginBottom":"6px"
                                                }
                                            ),
                                            dcc.Dropdown(
                                                id="gametag",
                                                options=[
                                                    {"label": x, "value": x}
                                                    for x in sorted(
                                                        df["Game Tags"]
                                                        .dropna()
                                                        .unique()
                                                    )
                                                ],
                                                multi=True
                                            )
                                        ]
                                    ),

                                    html.Div(
                                        [
                                            html.P(
                                                "Period Tags",
                                                style={
                                                    "color":"#f5f5f5",
                                                    "fontSize":"10px",
                                                    "marginBottom":"6px"
                                                }
                                            ),
                                            dcc.Dropdown(
                                                id="periodtag",
                                                options=[
                                                    {"label": x, "value": x}
                                                    for x in sorted(
                                                        df["Period Tags"]
                                                        .dropna()
                                                        .unique()
                                                    )
                                                ],
                                                multi=True
                                            )
                                        ]
                                    )
                                ],
                                className="filter-card",
                                style={
                                    "padding": "16px"
                                }
                            )
                        ],
                        className="filter-card",
                        style={
                            "padding": "16px"
                        }
                    )
                ],
                className="sidebar-panel sidebar-right-panel",
                style={
                    "flex": "0 0 auto",
                    "width": "180px",
                    "minWidth": "180px",
                    "padding": "24px",
                    "backgroundColor": "#011c24",
                    "borderRadius": "24px",
                    "border": "1px solid rgba(137, 188, 239, 0.16)",
                    "boxShadow": "0 18px 48px rgba(0,0,0,0.35)",
                    "position": "relative",
                    "top": "10px",
                    "maxHeight": "200vh",
                    "overflowY": "auto"
                }
            )
        ],
        style={
            "display": "flex",
            "alignItems": "flex-start",
            "gap": "18px",
            "width": "100%",
            "maxHeight": "calc(100vh - 100px)"
        }
    )
],
style={
    "color": "#ffffff",
    "padding": "28px 24px 36px",
    "fontFamily": "'Clash Display Semibold', 'Segoe UI', sans-serif",
    "margin": "0",
    "maxHeight": "calc(100vh - 100px)",
    "overflow": "auto"
}
)

@app.callback(
    Output("actividad-fecha-container","style"),
    Input("tabs","value")
)
def toggle_actividad_fecha(tab):
    if tab in ["actividad", "actividad_comparativa", "actividad_promedios"]:
        return {"display": "block", "marginTop": "10px"}
    return {"display": "none"}

@app.callback(
    Output("contenido-tab", "children"),
    Input("tabs", "value"),
    Input("categoria", "value"),
    Input("metrica", "value"),
    Input("referencia", "value"),
    Input("jugador", "value"),
    Input("athlete", "value"),
    Input("gametag", "value"),
    Input("periodtag", "value"),
    Input("fecha-actividad", "date")
)
def actualizar_tab(tab, categorias, metricas, referencia, jugadores, athlete, gametags, periodtags, fecha_actividad):
    dff = df.copy()
    # Filtros dinámicos
    if categorias: dff = dff[dff["Category"].isin(categorias)]
    if jugadores: dff = dff[dff["Player Name"].isin(jugadores)]
    if athlete: dff = dff[dff["Athlete Tags"].isin(athlete)]
    if gametags: dff = dff[dff["Game Tags"].isin(gametags)]
    if periodtags: dff = dff[dff["Period Tags"].isin(periodtags)]

    # Defaults
    metricas = metricas or ["Distance"]
    metricas = [m for m in metricas if m in dff.columns]
    referencia = referencia or "Category"
    fecha_dt = pd.to_datetime(fecha_actividad).normalize() if fecha_actividad else dff["Date"].max().normalize()
    
    # COMPARATIVAS
    if tab == "comparativas":
        fig = build_comparativas(dff, categorias, metricas, referencia)
        return html.Div([
            dcc.Graph(
                className="tab-graph",
                figure=fig,
                style={"width":"100%", "height":"100%"}
            )
        ], style={
            "border":"1px solid rgba(137,188,239,0.18)",
            "borderRadius":"18px",
            "overflow":"hidden",
            "background":"#0b0c0e",
            "boxShadow":(
                "0 0 20px rgba(72,247,136,0.10), "
                "0 0 50px rgba(137,188,239,0.08), "
                "0 18px 40px rgba(0,0,0,0.35)"
            ),
            "padding":"10px",
            "minWidth":"1000px",
            "width":"85%",
            "maxWidth":"1000px"
        })

        # ACTIVIDAD POR JUGADOR
    elif tab == "actividad":
            fecha_dt = pd.to_datetime(fecha_actividad).normalize() if fecha_actividad else dff["Date"].max().normalize()
            dff_fecha = dff[dff["Date"].dt.normalize() == fecha_dt]

            columnas_requeridas = [
                "Player Name","Accel + Decel Efforts","Accel + Decel Efforts Per Minute",
                "Distance","Player Load","Max Velocity","Meterage Per Minute",
                "Player Load Per Minute","Sprint Distance","Sprint Efforts","Sprint Dist Per Min",
                "High Speed Distance","High Speed Efforts","High Speed Distance Per Minute","Impacts"
            ]

            columnas_presentes = [c for c in columnas_requeridas if c in dff_fecha.columns]
            columnas_actividad = [{"name": c, "id": c} for c in columnas_presentes]

            estilos_condicionales = []
            for col in columnas_presentes:
                if col != "Player Name" and dff_fecha[col].dtype in ['float64','int64']:
                    max_val = dff_fecha[col].max()
                    min_val = dff_fecha[col].min()
                    rango = max_val - min_val if max_val != min_val else 1

                    estilos_condicionales.append({
                        "if": {"filter_query": f"{{{col}}} >= {max_val * 0.8}", "column_id": col},
                        "backgroundColor": "#017351","color": "white"
                    })
                    estilos_condicionales.append({
                        "if": {"filter_query": f"{{{col}}} >= {min_val + rango * 0.5} && {{{col}}} < {max_val * 0.8}", "column_id": col},
                        "backgroundColor": "#F4C95D","color": "black"
                    })
                    estilos_condicionales.append({
                        "if": {"filter_query": f"{{{col}}} < {min_val + rango * 0.5}", "column_id": col},
                        "backgroundColor": "#A40A1C","color": "white"
                    })

            return html.Div([
                html.H3("Actividad por Jugador", style={"color":"white","textAlign":"center","marginBottom":"10px",
                                                        "fontFamily":"'Clash Display Semibold', 'Helvetica Neue'","fontWeight":"600"}),
                html.H4(fecha_dt.strftime("%d/%m/%Y"), style={"color":"#a3e3d0","textAlign":"center","marginBottom":"15px",
                                                            "fontFamily":"'Clash Display Semibold', 'Helvetica Neue'","fontWeight":"600"}),
                dcc.Loading(
                    html.Div(
                        className="tab-table-wrapper",
                        children=[
                            dash_table.DataTable(
                                data=dff_fecha[columnas_presentes].to_dict("records") if columnas_presentes else [],
                                columns=columnas_actividad,
                                filter_action="none",
                                sort_action="native",
                                fixed_columns={"headers": True, "data": 1},
                                page_size=20,
                                style_table={"overflowX":"auto","minWidth":"100%","border":"1px solid rgba(137,188,239,0.18)",
                                            "boxShadow":"0 18px 40px rgba(0,0,0,0.25)"},
                                style_header={"backgroundColor":"#000000",
                                             "color":"white","fontWeight":"bold","position":"sticky","top":0,
                                              "textOverflow": "ellipsis"},
                                style_cell={  "backgroundColor":"#1a1a1a","color":"white",
                                                            "fontSize":"11px","textAlign":"center",
                                                            "minWidth":"100px",
                                                            "width":"100px",
                                                            "maxWidth":"100px",
                                                            "height":"15px",
                                                            "whiteSpace":"normal"
                                                                    },
                                style_data_conditional=estilos_condicionales
                            )
                        ]
                    )
                )
            ], style={
                    "padding": "22px",
                    "background": "#0b0c0e",
                    "border": "1px solid rgba(137,188,239,0.18)",
                    "borderRadius": "24px",
                    "boxShadow": "0 18px 40px rgba(0,0,0,0.25)",
                    "minWidth": "1000px",
                    "width": "1000px",
                    "maxWidth": "1000px"
                })
                        
  
# ACTIVIDAD COMPARATIVA INDIVIDUAL
    elif tab == "actividad_comparativa":
        fecha_dt = pd.to_datetime(fecha_actividad).normalize() if fecha_actividad else dff["Date"].max().normalize()
        dff_fecha = dff[dff["Date"].dt.normalize() == fecha_dt]

        metricas_base = [
            "Distance","Meterage Per Minute","Player Load","Player Load Per Minute",
            "Max Velocity","Accel + Decel Efforts","Accel + Decel Efforts Per Minute",
            "High Speed Distance","High Speed Distance Per Minute","High Speed Efforts",
            "Sprint Distance","Sprint Dist Per Min","Sprint Efforts","Impacts"
        ]

        resumen_fecha = dff_fecha.groupby("Player Name")[metricas_base].sum().reset_index()
        dff_acumulado = dff[dff["Date"].dt.normalize() <= fecha_dt]
        promedio_jugador = dff_acumulado.groupby("Player Name")[metricas_base].mean().reset_index()
        promedio_jugador = promedio_jugador.rename(columns={m: f"{m} Prom" for m in metricas_base})

        tabla_comparativa = resumen_fecha.merge(promedio_jugador, on="Player Name", how="left").fillna(0)

        columnas_comparativa = [{"name": "Player Name", "id": "Player Name"}]
        estilos_condicionales = []

        for m in metricas_base:
            columnas_comparativa.append({"name": m, "id": m, "type": "numeric", "format": {"specifier": ".2f"}})
            columnas_comparativa.append({"name": f"{m} Prom", "id": f"{m} Prom", "type": "numeric", "format": {"specifier": ".2f"}})

            # Calcular umbrales en Python
            for _, row in tabla_comparativa.iterrows():
                prom_val = row[f"{m} Prom"]
                if prom_val > 0:
                    umbral_alto = 1.3 * prom_val
                    umbral_bajo = 0.8 * prom_val

                    estilos_condicionales.append({
                        "if": {"filter_query": f"{{{m}}} > {umbral_alto}", "column_id": m},
                        "backgroundColor": "#017351", "color": "white", "lineHeight": "15px"
                    })
                    estilos_condicionales.append({
                        "if": {"filter_query": f"{{{m}}} >= {umbral_bajo} && {{{m}}} <= {umbral_alto}", "column_id": m},
                        "backgroundColor": "#e6c200", "color": "black", "lineHeight": "15px"
                    })
                    estilos_condicionales.append({
                        "if": {"filter_query": f"{{{m}}} < {umbral_bajo}", "column_id": m},
                        "backgroundColor": "#b22222", "color": "white", "lineHeight": "15px"
                    })

            # Diferenciar columna Prom
            estilos_condicionales.append({
                "if": {"column_id": f"{m} Prom"},
                "backgroundColor": "#2f2f2f", "color": "#d0d0d0", "fontWeight": "bold", "lineHeight": "15px"
            })

        return html.Div([
            html.H4("Comparativo última ACTIVIDAD vs PROMEDIO",
                    style={"color":"white","textAlign":"center","marginBottom":"10px",
                                                        "fontFamily":"'Clash Display Semibold', 'Helvetica Neue'","fontWeight":"600"}),
            dcc.Loading(
                dash_table.DataTable(
                    data=tabla_comparativa.to_dict("records"),
                    columns=columnas_comparativa,
                    filter_action="none",
                    sort_action="native",
                    fixed_columns={"headers": True, "data": 1},
                    page_size=20,
                    style_table={"overflowX": "auto", "minWidth": "100%",
                                "border": "1px solid rgba(137,188,239,0.18)",
                                "boxShadow": "0 18px 40px rgba(0,0,0,0.25)"},
                    style_header={"backgroundColor": "#000000", "color": "white",
                                "height":"25px","maxHeight":"25px",
                                "minWidth":"100px","width":"100px","maxWidth":"100px",
                                "fontWeight": "bold", "position": "sticky", "top": 0,
                                "textOverflow": "ellipsis"},
                    style_cell={ "backgroundColor":"#1a1a1a","color":"white",
                                                "fontSize":"11px","textAlign":"center",
                                                "minWidth":"100px",
                                                "width":"100px",
                                                "maxWidth":"100px",
                                                "height":"15px",
                                                "whiteSpace":"normal"            },
                    style_data_conditional=estilos_condicionales
                )
            )
        ], style={"padding": "22px", "background": "#0b0c0e",
                "border": "1px solid rgba(137,188,239,0.18)", "borderRadius": "24px",
                "boxShadow": "0 18px 40px rgba(0,0,0,0.25)",
                "minWidth": "1000px", "overflowX": "auto",
                "maxWidth": "1000px",
                "maxHeight": "600px", "overflowY": "auto"
                })

    elif tab == "actividad_promedios":
        fecha_dt = pd.to_datetime(fecha_actividad).normalize() if fecha_actividad else dff["Date"].max().normalize()
        dff_fecha = dff[dff["Date"].dt.normalize() == fecha_dt]

        metricas_promedios_validas = [m for m in metricas_promedios if m in dff_fecha.columns]
        resumen = dff_fecha[metricas_promedios_validas].mean().round(2).to_dict() if not dff_fecha.empty else {}

        categoria_text = summarize_items(categorias, max_items=10, default="Todas")
        gametag_text = summarize_items(gametags, max_items=10, default="Todos")
        periodtag_text = summarize_items(periodtags, max_items=10, default="Todos")
        fecha_text = fecha_dt.strftime("%d/%m/%Y")
        grafico_titulo = (
            f"Actividad {fecha_text}  |  Dinámica: {gametag_text}  |  Microciclo: {periodtag_text}  |  Plantel: {categoria_text}"
        )

        cards = []
        for m in metricas_promedios_validas:
            valor = resumen.get(m, 0)
            cards.append(
                html.Div([
                    html.Div(m, style={"color": "#a3e3d0",
            "fontSize": "12px",
            "marginTop": "1px",
            "textAlign": "left",
            "lineHeight": "1.2"}),
                    html.Div(f"{valor:.2f}", style={"color": "#edf1f2",
            "fontSize": "30px",
            "fontWeight": "700",
            "letterSpacing": "1px",
            "marginTop": "2px",  
            "marginBottom": "2px",
            "textAlign": "left",   
            "lineHeight": "1.2"})
                ], style={
                    "padding": "20px",
                    "background": "#0b0c0e",
                    "border": "1px solid rgba(137,188,239,0.18)",
                    "borderRadius": "20px",
                    "boxShadow": "0 18px 40px rgba(0,0,0,0.25)",
                    "minWidth": "180px",
                    "maxWidth": "180px",
                    "maxHeight": "50px",
                    "flex": "1",
                    "flexWrap": "wrap",
                    "gap": "16px",
                    "padding": "16px"
                })
            )

        return html.Div([
            html.Div([
                html.H3(grafico_titulo, style={"color":"white","textAlign":"center","marginTop":"20px", 
                    "fontFamily":"'Clash Display Semibold', 'Helvetica Neue'","fontWeight":"600"}),
                html.Div([
                    html.Div([html.Div("Fecha de actividad", style={"color": "#a3e3d0", "fontSize": "12px", "marginBottom": "4px"}),
                              html.Div(fecha_text, style={"color": "#edf1f2", "fontSize": "14px", "fontWeight": "600"})],
                             style={"minWidth": "180px", "maxWidth": "180px", "padding": "16px", "background": "#071016", "borderRadius": "18px", "border": "1px solid rgba(137,188,239,0.18)"}),
                    html.Div([html.Div("Game Tags", style={"color": "#a3e3d0", "fontSize": "12px", "marginBottom": "4px"}),
                              html.Div(gametag_text, style={"color": "#edf1f2", "fontSize": "14px", "fontWeight": "600"})],
                             style={"minWidth": "180px", "maxWidth": "180px", "padding": "16px", "background": "#071016", "borderRadius": "18px", "border": "1px solid rgba(137,188,239,0.18)"}),
                    html.Div([html.Div("Period Tags", style={"color": "#a3e3d0", "fontSize": "12px", "marginBottom": "4px"}),
                              html.Div(periodtag_text, style={"color": "#edf1f2", "fontSize": "14px", "fontWeight": "600"})],
                             style={"minWidth": "180px", "maxWidth": "180px", "padding": "16px", "background": "#071016", "borderRadius": "18px", "border": "1px solid rgba(137,188,239,0.18)"}),
                    html.Div([html.Div("Category", style={"color": "#a3e3d0", "fontSize": "12px", "marginBottom": "4px"}),
                              html.Div(categoria_text, style={"color": "#edf1f2", "fontSize": "14px", "fontWeight": "600"})],
                             style={"minWidth": "180px", "maxWidth": "180px", "padding": "16px", "background": "#071016", "borderRadius": "18px", "border": "1px solid rgba(137,188,239,0.18)"})
                ], style={"display": "grid", "gridTemplateColumns": "repeat(auto-fit, minmax(180px, 1fr))", "gap": "16px", "marginBottom": "24px"})
            ], style={"marginBottom": "10px", "padding": "10px 10px"}),
            html.Div(
                cards if cards else [html.Div("No hay datos para la fecha seleccionada.", style={"color": "#edf1f2", "textAlign": "center", "padding": "24px"})],
                style={"minWidth":"180px", "display": "grid", "gridTemplateColumns": "repeat(auto-fit, minmax(180px, 1fr))", "flexWrap":"wrap" ,"gap": "16px", "marginBottom":"24px"}
            )
        ], style={"padding": "20px", "background": "linear-gradient(145deg, #0b0c0e, #1a1c1f)",
                    "border": "1px solid rgba(137,188,239,0.25)", "borderRadius": "28px",
                    "boxShadow": "0 12px 30px rgba(0,0,0,0.35)", "margin": "20px auto", "maxWidth": "1100px"})


# ACWR
    
    elif tab=="acwr":
        metricas_acwr = list(dict.fromkeys([

            "Distance",
            "Player Load",
            "Acceleration Efforts",
            "Sprint Distance",
            "High Speed Distance",
            "Sprint Efforts",
            "High Speed Efforts",
            "Impacts"

        ]))

        dff["Player Name"] = (
            dff["Player Name"]
            .astype(str)
            .str.strip()
        )

        ultimos21 = (
            dff["Date"].max()
            - pd.Timedelta(days=21)
        )

        ultimos7 = (
            dff["Date"].max()
            - pd.Timedelta(days=7)
        )

        df21 = dff[
            dff["Date"] >= ultimos21
        ]

        df7 = dff[
            dff["Date"] >= ultimos7
        ]

        cronica = (

            df21
            .groupby("Player Name")[metricas_acwr]
            .mean()
            .reset_index()

        )

        aguda = (

            df7
            .groupby("Player Name")[metricas_acwr]
            .mean()
            .reset_index()

        )

        tabla = cronica.merge(

            aguda,
            on="Player Name",
            how="outer",
            suffixes=("_21", "_7")

        )

        tabla = tabla.loc[
            :,
            ~tabla.columns.duplicated()
        ]

        for m in metricas_acwr:

            tabla[f"{m}_ACWR"] = (

                tabla[f"{m}_7"] /
                tabla[f"{m}_21"]

            ).round(2)

        ratio_columns = [

            f"{m}_ACWR"

            for m in metricas_acwr

        ]

        tabla = tabla[
            ["Player Name"] + ratio_columns
        ].fillna(0)

        return html.Div([

            html.H3(

                "ACWR - Últimos 7 días vs 21 días",

                style={
                "color":"white","textAlign":"center","marginBottom":"10px",
                "fontFamily":"'Clash Display Semibold', 'Helvetica Neue'","fontWeight":"600"
            }),

            dcc.Loading(

                html.Div(
                    className="tab-table-wrapper",
                    children=[
                        dash_table.DataTable(

                            data=tabla.to_dict("records"),

                            columns=[

                        {
                            "name":"Player Name",
                            "id":"Player Name"
                        }

                    ] + [

                        {
                            "name":col,
                            "id":col,
                            "type":"numeric",
                            "format":{"specifier":".2f"}
                        }

                        for col in ratio_columns

                    ],
                        filter_action="none",
                        sort_action="native",
                        fixed_columns={"headers": True, "data": 1},
                        style_table={
                            "overflowX": "auto",
                            "minWidth": "100%"
                        },
                        style_header={
                            "backgroundColor": "#0d1620",
                            "color": "#edf1f2",
                            "fontWeight": "bold",
                            "borderBottom": "1px solid rgba(137,188,239,0.2)"
                        },
                        style_cell={
                            "backgroundColor":"#1a1a1a","color":"white",
                            "fontSize":"11px","textAlign":"center",
                            "minWidth":"100px",
                            "width":"100px",
                            "maxWidth":"100px",
                            "height":"15px",
                            "whiteSpace":"normal"
                        },
                        style_data_conditional=[
                            {"lineHeight": "15px",
                                "if": {
                                    "filter_query": "{Distance_ACWR} < 0.8",
                                    "column_id": "Distance_ACWR"
                                },
                                "backgroundColor": "#A40A1C",
                                "color": "white"
                            },
                            {
                                "if": {
                                    "filter_query": "{Distance_ACWR} >= 0.8 && {Distance_ACWR} <= 1.3",
                                    "column_id": "Distance_ACWR"
                                },
                                "backgroundColor": "#017351",
                                "color": "white"
                            },
                            {
                                "if": {
                                    "filter_query": "{Distance_ACWR} > 1.3",
                                    "column_id": "Distance_ACWR"
                                },
                                "backgroundColor": "#ff5e5e",
                                "color": "black"
                            },
                            {
                                "if": {
                                    "filter_query": "{Player Load_ACWR} < 0.8",
                                    "column_id": "Player Load_ACWR"
                                },
                                "backgroundColor": "#A40A1C",
                                "color": "white"
                            },
                            {
                                "if": {
                                    "filter_query": "{Player Load_ACWR} >= 0.8 && {Player Load_ACWR} <= 1.3",
                                    "column_id": "Player Load_ACWR"
                                },
                                "backgroundColor": "#017351",
                                "color": "white"
                            },
                            {
                                "if": {
                                    "filter_query": "{Player Load_ACWR} > 1.3",
                                    "column_id": "Player Load_ACWR"
                                },
                                "backgroundColor": "#ff5e5e",
                                "color": "black"
                            },
                            {
                                "if": {
                                    "filter_query": "{Acceleration Efforts_ACWR} < 0.8",
                                    "column_id": "Acceleration Efforts_ACWR"
                                },
                                "backgroundColor": "#A40A1C",
                                "color": "white"
                            },
                            {
                                "if": {
                                    "filter_query": "{Acceleration Efforts_ACWR} >= 0.8 && {Acceleration Efforts_ACWR} <= 1.3",
                                    "column_id": "Acceleration Efforts_ACWR"
                                },
                                "backgroundColor": "#017351",
                                "color": "white"
                            },
                            {
                                "if": {
                                    "filter_query": "{Acceleration Efforts_ACWR} > 1.3",
                                    "column_id": "Acceleration Efforts_ACWR"
                                },
                                "backgroundColor": "#ff5e5e",
                                "color": "black"
                            },
                            {
                                "if": {
                                    "filter_query": "{Sprint Distance_ACWR} < 0.8",
                                    "column_id": "Sprint Distance_ACWR"
                                },
                                "backgroundColor": "#A40A1C",
                                "color": "white"
                            },
                            {
                                "if": {
                                    "filter_query": "{Sprint Distance_ACWR} >= 0.8 && {Sprint Distance_ACWR} <= 1.3",
                                    "column_id": "Sprint Distance_ACWR"
                                },
                                "backgroundColor": "#017351",
                                "color": "white"
                            },
                            {
                                "if": {
                                    "filter_query": "{Sprint Distance_ACWR} > 1.3",
                                    "column_id": "Sprint Distance_ACWR"
                                },
                                "backgroundColor": "#ff5e5e",
                                "color": "black"
                            },
                            {
                                "if": {
                                    "filter_query": "{High Speed Distance_ACWR} < 0.8",
                                    "column_id": "High Speed Distance_ACWR"
                                },
                                "backgroundColor": "#A40A1C",
                                "color": "white"
                            },
                            {
                                "if": {
                                    "filter_query": "{High Speed Distance_ACWR} >= 0.8 && {High Speed Distance_ACWR} <= 1.3",
                                    "column_id": "High Speed Distance_ACWR"
                                },
                                "backgroundColor": "#017351",
                                "color": "white"
                            },
                            {
                                "if": {
                                    "filter_query": "{High Speed Distance_ACWR} > 1.3",
                                    "column_id": "High Speed Distance_ACWR"
                                },
                                "backgroundColor": "#ff5e5e",
                                "color": "black"
                            },
                            {
                                "if": {
                                    "filter_query": "{Sprint Efforts_ACWR} < 0.8",
                                    "column_id": "Sprint Efforts_ACWR"
                                },
                                "backgroundColor": "#A40A1C",
                                "color": "white"
                            },
                            {
                                "if": {
                                    "filter_query": "{Sprint Efforts_ACWR} >= 0.8 && {Sprint Efforts_ACWR} <= 1.3",
                                    "column_id": "Sprint Efforts_ACWR"
                                },
                                "backgroundColor": "#017351",
                                "color": "white"
                            },
                            {
                                "if": {
                                    "filter_query": "{Sprint Efforts_ACWR} > 1.3",
                                    "column_id": "Sprint Efforts_ACWR"
                                },
                                "backgroundColor": "#ff5e5e",
                                "color": "black"
                            },
                            {
                                "if": {
                                    "filter_query": "{High Speed Efforts_ACWR} < 0.8",
                                    "column_id": "High Speed Efforts_ACWR"
                                },
                                "backgroundColor": "#A40A1C",
                                "color": "white"
                            },
                            {
                                "if": {
                                    "filter_query": "{High Speed Efforts_ACWR} >= 0.8 && {High Speed Efforts_ACWR} <= 1.3",
                                    "column_id": "High Speed Efforts_ACWR"
                                },
                                "backgroundColor": "#017351",
                                "color": "white"
                            },
                            {
                                "if": {
                                    "filter_query": "{High Speed Efforts_ACWR} > 1.3",
                                    "column_id": "High Speed Efforts_ACWR"
                                },
                                "backgroundColor": "#ff5e5e",
                                "color": "white"
                            },
                            {
                                "if": {
                                    "filter_query": "{Impacts_ACWR} < 0.8",
                                    "column_id": "Impacts_ACWR"
                                },
                                "backgroundColor": "#A40A1C",
                                "color": "white"
                            },
                            {
                                "if": {
                                    "filter_query": "{Impacts_ACWR} >= 0.8 && {Impacts_ACWR} <= 1.3",
                                    "column_id": "Impacts_ACWR"
                                },
                                "backgroundColor": "#017351",
                                "color": "white"
                            },
                            {
                                "if": {
                                    "filter_query": "{Impacts_ACWR} > 1.3",
                                    "column_id": "Impacts_ACWR"
                                },
                                "backgroundColor": "#ff5e5e",
                                "color": "white",
                                
                            }
                        ],
                        page_size=20 
                        
                            )
                        ]
                    )
                )
            ],style={
            "padding": "18px",
            "background": "#0b0c0e",
            "border": "1px solid rgba(137,188,239,0.18)",
            "borderRadius": "24px",
            "boxShadow": "0 18px 40px rgba(0,0,0,0.25)",
            "minHeight": "600px",
            "minWidth": "1000px",
            "width": "100%",
            "maxWidth": "1000px"
})
    
        # CRONOLÓGICO
    elif tab == "cronologico": 
        cronologico = pd.melt(
        dff,
            id_vars=["Date", "Category"],
            value_vars=metricas,
            var_name="Métrica",
            value_name="Valor"
            )  
        fig = px.scatter(
            cronologico,
            x="Date",
            y="Valor",
            color="Category",
            symbol="Métrica",
            color_discrete_sequence=["#edf1f2", "#f1a3fd", "#a3e3d0", "#89bcef", "#48f788", "#f96e83"],
            template="plotly_dark"
        )

        fig.update_traces(
            marker=dict(size=10, line=dict(width=1, color="#ffffff")),
            selector=dict(mode="markers"),
            hoverlabel=dict(bgcolor="#011c24", font_size=12, font_color="#f5f5f5")
        )

        fig.update_layout(
            title={
                "text": 'Cronológico',
                "font": {
                    "color": "#f5f5f5",
                    "family": "'Clash Display Semibold', 'Helvetica Neue'",
                    "size": 22
                }
            },
            paper_bgcolor="#0b0c0e",
            plot_bgcolor="#0b0c0e",
            font={"color": "#f5f5f5"},
            legend=dict(
                bgcolor="rgba(11,12,14,0.75)",
                bordercolor="#89bcef",
                borderwidth=1
            ))
            
        if LOGO_BASE64:
            fig.add_layout_image(
                dict(
                    source="data:image/png;base64," + LOGO_BASE64,
                    xref="paper",
                    yref="paper",
                    x=0.99,
                    y=0.01,
                    xanchor="right",
                    yanchor="bottom",
                    sizex=0.12,
                    sizey=0.10,
                    opacity=0.7,
                    layer="above"
                )
            )

        fig.update_xaxes(
            tickformat="%d/%m/%Y",
            showgrid=True,
            gridcolor="rgba(137,188,239,0.18)",
            zerolinecolor="rgba(255,255,255,0.08)",
            linecolor="#89bcef",
            tickfont_color="#f5f5f5",
            title_font_color="#a3e3d0"
        )

        fig.update_yaxes(
            showgrid=True,
            gridcolor="rgba(137,188,239,0.18)",
            zerolinecolor="rgba(255,255,255,0.08)",
            linecolor="#89bcef",
            tickfont_color="#f5f5f5",
            title_font_color="#a3e3d0"
        )

        return html.Div(
            dcc.Graph(
                className="tab-graph",
                figure=fig,
                style={"width": "100%", "height": "100%"}
            ),
                style={"padding": "22px", "background": "#0b0c0e",
                    "border": "1px solid rgba(137,188,239,0.18)", "borderRadius": "24px",
                    "boxShadow": "0 18px 40px rgba(0,0,0,0.25)",
                    "minHeight": "600px",
                                "minWidth": "1000px",
                                "width": "85%",
                                "maxWidth": "1000px"})

    elif tab == "informe":
        return html.Div([
            html.H2("Informe de Actividad", style={"color":"#edf1f2","textAlign":"center","marginBottom":"20px"}),
            html.Div([
                html.Div([
                    html.Label("Título del informe", style={"color":"#a3e3d0","marginBottom":"6px"}),
                    dcc.Input(id="report_title", type="text", placeholder="Título personalizado del informe (opcional)", value="", style={"width":"100%","padding":"10px","borderRadius":"12px","border":"1px solid rgba(137,188,239,0.18)","background":"#071016","color":"#edf1f2"}),
                    html.Div("Deja el título en blanco para generar uno automático.", style={"color":"#dcdcdc","fontSize":"11px","marginTop":"6px"})
                ], style={"flex":"1"}),
                html.Div([
                    html.Label("Creado por", style={"color":"#a3e3d0","marginBottom":"6px"}),
                    dcc.Input(id="report_author", type="text", placeholder="Nombre del creador", value="", style={"width":"100%","padding":"10px","borderRadius":"12px","border":"1px solid rgba(137,188,239,0.18)","background":"#071016","color":"#edf1f2"})
                ], style={"flex":"1","minWidth":"260px","marginLeft":"16px"}),
                html.Div([
                    html.Label("Fecha de actividad", style={"color":"#a3e3d0","marginBottom":"6px"}),
                    dcc.DatePickerSingle(
                        id="report_fecha_actividad",
                        date=fecha_max.date(),
                        min_date_allowed=df["Date"].min().date(),
                        max_date_allowed=df["Date"].max().date(),
                        display_format="DD/MM/YYYY",
                        style={
                            "width": "100%",
                            "backgroundColor": "#011c24",
                            "color": "#f5f5f5",
                            "border": "1px solid rgba(137,188,239,0.18)",
                            "borderRadius": "12px",
                            "padding": "10px"
                        }
                    ),
                    html.Div("Selecciona la fecha de actividad para construir el informe.", style={"color":"#dcdcdc","fontSize":"11px","marginTop":"6px"})
                ], style={"flex":"1","minWidth":"260px","marginLeft":"16px"})
            ], style={"display":"flex","gap":"16px","marginBottom":"24px","flexWrap":"wrap"}),
            html.Div([
                html.Label("Secciones del informe", style={"color":"#a3e3d0","marginBottom":"8px"}),
                dcc.Checklist(
                    id="report_sections",
                    options=[
                        {"label": section_title(v), "value": v} for v in [
                            "actividad",
                            "actividad_comparativa",
                            "actividad_promedios",
                            "acwr",
                            "plyr_vs_plyr",
                            "comparativas",
                            "cronologico"
                        ]
                    ],
                    value=["actividad", "actividad_promedios", "acwr"],
                    labelStyle={"display":"block", "color":"#edf1f2", "marginBottom":"6px"},
                    inputStyle={"marginRight":"8px"}
                )
            ], style={"marginBottom":"28px","padding":"20px","background":"#071016","borderRadius":"18px","border":"1px solid rgba(137,188,239,0.18)"}),
            html.Div([
                html.Div([
                    html.Label("Actividad", style={"color":"#a3e3d0","marginBottom":"6px"}),
                    dcc.Textarea(id="report_text_actividad", placeholder="Describe la sección de Actividad. Hasta 500 palabras.", style={"width":"100%","height":"140px","borderRadius":"16px","border":"1px solid rgba(137,188,239,0.18)","background":"#071016","color":"#edf1f2"})
                ], style={"flex":"1","minWidth":"280px","marginBottom":"16px"}),
                html.Div([
                    html.Label("Actividad Comparativa Individual", style={"color":"#a3e3d0","marginBottom":"6px"}),
                    dcc.Textarea(id="report_text_actividad_comparativa", placeholder="Describe la sección de Actividad Comparativa Individual. Hasta 500 palabras.", style={"width":"100%","height":"140px","borderRadius":"16px","border":"1px solid rgba(137,188,239,0.18)","background":"#071016","color":"#edf1f2"})
                ], style={"flex":"1","minWidth":"280px","marginBottom":"16px"})
            ], style={"display":"flex","flexWrap":"wrap","gap":"16px","marginBottom":"16px"}),
            html.Div([
                html.Div([
                    html.Label("Actividad/Promedios", style={"color":"#a3e3d0","marginBottom":"6px"}),
                    dcc.Textarea(id="report_text_actividad_promedios", placeholder="Describe la sección de Actividad/Promedios. Hasta 500 palabras.", style={"width":"100%","height":"140px","borderRadius":"16px","border":"1px solid rgba(137,188,239,0.18)","background":"#071016","color":"#edf1f2"})
                ], style={"flex":"1","minWidth":"280px","marginBottom":"16px"}),
                html.Div([
                    html.Label("ACWR", style={"color":"#a3e3d0","marginBottom":"6px"}),
                    dcc.Textarea(id="report_text_acwr", placeholder="Describe la sección de ACWR. Hasta 500 palabras.", style={"width":"100%","height":"140px","borderRadius":"16px","border":"1px solid rgba(137,188,239,0.18)","background":"#071016","color":"#edf1f2"})
                ], style={"flex":"1","minWidth":"280px","marginBottom":"16px"})
            ], style={"display":"flex","flexWrap":"wrap","gap":"16px","marginBottom":"16px"}),
            html.Div([
                html.Div([
                    html.Label("PLYR vs PLYR", style={"color":"#a3e3d0","marginBottom":"6px"}),
                    dcc.Textarea(id="report_text_plyr_vs_plyr", placeholder="Describe la sección de PLYR vs PLYR. Hasta 500 palabras.", style={"width":"100%","height":"140px","borderRadius":"16px","border":"1px solid rgba(137,188,239,0.18)","background":"#071016","color":"#edf1f2"})
                ], style={"flex":"1","minWidth":"280px","marginBottom":"16px"}),
                html.Div([
                    html.Label("Comparativo", style={"color":"#a3e3d0","marginBottom":"6px"}),
                    dcc.Textarea(id="report_text_comparativas", placeholder="Describe la sección de Comparativo. Hasta 500 palabras.", style={"width":"100%","height":"140px","borderRadius":"16px","border":"1px solid rgba(137,188,239,0.18)","background":"#071016","color":"#edf1f2"})
                ], style={"flex":"1","minWidth":"280px","marginBottom":"16px"})
            ], style={"display":"flex","flexWrap":"wrap","gap":"16px","marginBottom":"16px"}),
            html.Div([
                html.Div([
                    html.Label("Cronológico", style={"color":"#a3e3d0","marginBottom":"6px"}),
                    dcc.Textarea(id="report_text_cronologico", placeholder="Describe la sección Cronológico. Hasta 500 palabras.", style={"width":"100%","height":"140px","borderRadius":"16px","border":"1px solid rgba(137,188,239,0.18)","background":"#071016","color":"#edf1f2"})
                ], style={"flex":"1","minWidth":"280px","marginBottom":"16px"})
            ], style={"display":"flex","flexWrap":"wrap","gap":"16px","marginBottom":"24px"}),
            html.Div(id="report_figures_preview", style={"display":"grid","gridTemplateColumns":"repeat(auto-fit,minmax(320px,1fr))","gap":"20px","marginBottom":"24px"}),
            html.Div([
                html.Button("Generar PDF", id="generate_report", n_clicks=0, style={"width":"100%","padding":"16px","borderRadius":"18px","border":"none","background":"#89bcef","color":"#0b0c0e","fontWeight":"700","cursor":"pointer"})
            ], style={"maxWidth":"320px","margin":"0 auto"}, style_selected={"background":"#48f788","color":"#0b0c0e"}),
            html.Div("Al hacer clic se generará un PDF con secciones seleccionadas, texto y gráficos incrustados.", style={"color":"#dcdcdc","fontSize":"12px","textAlign":"center","marginTop":"12px"})
        ], style={"padding":"24px","background":"#0b0c0e","border":"1px solid rgba(137,188,239,0.18)","borderRadius":"28px","boxShadow":"0 18px 40px rgba(0,0,0,0.25)","maxWidth":"1180px","margin":"20px auto"})

    # ======================================================
# PLYR vs PLYR
# ======================================================

    elif tab == "plyr_vs_plyr":
        return html.Div([
            html.H3(
                "Comparativa Jugador vs Jugador", 
                style={
                    "color":"white",
                    "textAlign":"center",
                    "fontSize":"22px",
                    "marginBottom":"20px",
                    "fontFamily":"'Clash Display Semibold', 'Helvetica Neue'",
                    "fontWeight":"600"
                }
            ),

            # Dropdowns para elegir jugadores
            html.Div([
                dcc.Dropdown(
                    id="jugador_1",
                    options=[{"label": j, "value": j} for j in dff["Player Name"].unique()],
                    value=dff["Player Name"].unique()[0],
                    placeholder="Seleccionar Jugador 1",
                    style={"width":"45%","display":"inline-block","marginRight":"10px", "color":"#0d1620"}
                ),
                dcc.Dropdown(
                    id="jugador_2",
                    options=[{"label": j, "value": j} for j in dff["Player Name"].unique()],
                    value=dff["Player Name"].unique()[1] if len(dff["Player Name"].unique())>1 else None,
                    placeholder="Seleccionar Jugador 2",
                    style={"width":"45%","display":"inline-block","color":"#0d1620"}
                )
            ], style={"display":"flex","justifyContent":"center","marginBottom":"20px"}),

            # Filtros adicionales
            html.Div([
                dcc.Dropdown(
                    id="game_tags",
                    options=[{"label": g, "value": g} for g in dff["Game Tags"].unique()],
                    placeholder="Filtrar por Game Tag",
                    style={"width":"45%","display":"inline-block","marginRight":"10px","color":"#0d1620"}
                ),
                dcc.Dropdown(
                    id="period_tags",
                    options=[{"label": p, "value": p} for p in dff["Period Tags"].unique()],
                    placeholder="Filtrar por Period Tag",
                    style={"width":"45%","display":"inline-block", "color":"#0d1620"}
                )
            ], style={"display":"flex","justifyContent":"center","marginBottom":"20px"}),

            # Gráfico
            dcc.Graph(id="radar_chart", className="tab-graph", style={"height":"600px"})
        ], style={
            "padding":"28px",
            "background":"linear-gradient(145deg, #0b0c0e, #1a1c1f)",
            "border":"1px solid rgba(137,188,239,0.25)",
            "borderRadius":"28px",
            "boxShadow":"0 12px 30px rgba(0,0,0,0.35)",
            "margin":"20px auto",
            "maxWidth":"600px",
            "minWidth": "600px",
            "minHeight": "600px",
            "width": "100%"
        })

    else:
        return no_update


@app.callback(
    Output("gametag", "options"),
    Output("gametag", "value"),
    Output("periodtag", "options"),
    Output("periodtag", "value"),
    Input("fecha-actividad", "date"),
    Input("categoria", "value")
)
def actualizar_tags_por_fecha_categoria(fecha_actividad, categorias):
    dff = df.copy()
    if fecha_actividad:
        fecha_dt = pd.to_datetime(fecha_actividad).normalize()
        dff = dff[dff["Date"].dt.normalize() == fecha_dt]
    if categorias:
        dff = dff[dff["Category"].isin(categorias)]

    gametag_vals = sorted(dff["Game Tags"].dropna().unique())
    periodtag_vals = sorted(dff["Period Tags"].dropna().unique())

    gametag_options = [{"label": x, "value": x} for x in gametag_vals]
    periodtag_options = [{"label": x, "value": x} for x in periodtag_vals]

    return gametag_options, gametag_vals, periodtag_options, periodtag_vals

@app.callback(
    Output("report_figures_preview", "children"),
    Input("report_sections", "value"),
    Input("categoria", "value"),
    Input("report_fecha_actividad", "date")
)
def actualizar_vista_previa_informe(sections, categorias, fecha_actividad):
    if not sections:
        return html.Div(
            "Selecciona al menos una sección para ver las figuras en el informe.",
            style={
                "color": "#edf1f2",
                "textAlign": "center",
                "padding": "20px",
                "background": "#0b0c0e",
                "border": "1px solid rgba(137,188,239,0.18)",
                "borderRadius": "24px"
            }
        )

    dff = df.copy()
    if categorias:
        dff = dff[dff["Category"].isin(categorias)]

    fecha_dt = pd.to_datetime(fecha_actividad).normalize() if fecha_actividad else dff["Date"].max().normalize()
    if fecha_dt is not None:
        dff = dff[dff["Date"].dt.normalize() <= fecha_dt]

    preview_cards = []
    for section in sections:
        try:
            fig = build_section_report_fig(section, dff, fecha_dt, categorias)
        except Exception as e:
            logging.warning("Error construyendo figura para sección %s: %s", section, e)
            fig = None

        if fig is None or not getattr(fig, "data", None):
            card = html.Div(
                [
                    html.H4(section_title(section), style={"color": "#a3e3d0", "marginBottom": "10px"}),
                    html.Div(
                        "No hay figura disponible para esta sección con los filtros seleccionados.",
                        style={"color": "#edf1f2", "fontSize": "14px"}
                    )
                ],
                style={
                    "padding": "18px",
                    "background": "#0b0c0e",
                    "border": "1px solid rgba(137,188,239,0.18)",
                    "borderRadius": "24px",
                    "boxShadow": "0 18px 40px rgba(0,0,0,0.25)"
                }
            )
        else:
            card = html.Div(
                [
                    html.H4(section_title(section), style={"color": "#a3e3d0", "marginBottom": "10px"}),
                    dcc.Graph(
                        className="tab-graph",
                        figure=fig,
                        config={"displayModeBar": False, "responsive": True},
                        style={"height": "600px", "width": "100%"}
                    )
                ],
                style={
                    "padding": "18px",
                    "background": "#0b0c0e",
                    "border": "1px solid rgba(137,188,239,0.18)",
                    "borderRadius": "24px",
                    "boxShadow": "0 18px 40px rgba(0,0,0,0.25)"
                }
            )
        preview_cards.append(card)

    return preview_cards


@app.callback(
    Output("download-report", "data"),
    Input("generate_report", "n_clicks"),
    State("report_title", "value"),
    State("report_author", "value"),
    State("report_sections", "value"),
    State("report_text_actividad", "value"),
    State("report_text_actividad_comparativa", "value"),
    State("report_text_actividad_promedios", "value"),
    State("report_text_acwr", "value"),
    State("report_text_plyr_vs_plyr", "value"),
    State("report_text_comparativas", "value"),
    State("report_text_cronologico", "value"),
    State("categoria", "value"),
    State("report_fecha_actividad", "date"),
    State("download_format", "value"),
    prevent_initial_call=True
)
def generar_informe(
    n_clicks,
    title,
    author,
    sections,
    texto_actividad,
    texto_actividad_comparativa,
    texto_actividad_promedios,
    texto_acwr,
    texto_plyr_vs_plyr,
    texto_comparativas,
    texto_cronologico,
    categorias,
    fecha_actividad,
    download_format
):

    if not n_clicks:
        return no_update

    # Asegurar lista de categorías
    if categorias and isinstance(categorias, str):
        categorias = [categorias]

    title = (title or "").strip() or build_auto_report_title(
        categorias,
        fecha_actividad
    )

    author = (author or "").strip() or "Desconocido"

    fecha_text = (
        pd.to_datetime(fecha_actividad).strftime("%d/%m/%Y")
        if fecha_actividad
        else datetime.now().strftime("%d/%m/%Y")
    )

    selected_sections = sections or [
        "actividad",
        "actividad_promedios",
        "acwr"
    ]

    if not selected_sections:
        return no_update

    # Copia del dataframe
    dff = df.copy()

    # Garantizar datetime
    if "Date" in dff.columns:
        dff["Date"] = pd.to_datetime(
            dff["Date"],
            errors="coerce"
        )

    # Filtrar categorías
    if categorias:
        dff = dff[dff["Category"].isin(categorias)]

    # Fecha de corte
    if fecha_actividad:
        fecha_dt = pd.to_datetime(
            fecha_actividad,
            errors="coerce"
        )

        if pd.notna(fecha_dt):
            fecha_dt = fecha_dt.normalize()
    else:
        fecha_dt = (
            dff["Date"].max().normalize()
            if (
                "Date" in dff.columns
                and not dff.empty
                and dff["Date"].notna().any()
            )
            else None
        )

    if fecha_dt is not None and "Date" in dff.columns:
        dff = dff[
            dff["Date"].dt.normalize() <= fecha_dt
        ]

    section_texts = {
        "actividad": texto_actividad or "",
        "actividad_comparativa": texto_actividad_comparativa or "",
        "actividad_promedios": texto_actividad_promedios or "",
        "acwr": texto_acwr or "",
        "plyr_vs_plyr": texto_plyr_vs_plyr or "",
        "comparativas": texto_comparativas or "",
        "cronologico": texto_cronologico or ""
    }

    filters_parts = []

    if categorias:
        filters_parts.append(
            f"Categorías: {', '.join(categorias)}"
        )

    filters_parts.append(
        f"Fecha: {fecha_text}"
    )

    filters_text = " | ".join(filters_parts)

    report_sections = []
    image_bytes_for_png = []

    for section in selected_sections:

        if section == "informe":
            continue

        try:
            fig = build_section_report_fig(
                section,
                dff,
                fecha_dt,
                categorias
            )

            table_fig = build_section_report_table_fig(
                section,
                dff,
                fecha_dt,
                categorias
            )

        except Exception:
            logging.exception(
                f"Error generando figuras para {section}"
            )
            continue

        img_bytes = None
        table_bytes = None

        try:
            if fig and getattr(fig, "data", None):
                img_bytes = fig_to_png_bytes(
                    fig,
                    width=1600,
                    height=900,
                    scale=2
                )
        except Exception:
            logging.exception(
                f"Error exportando gráfico {section}"
            )

        try:
            if table_fig and getattr(table_fig, "data", None):
                table_bytes = fig_to_png_bytes(
                    table_fig,
                    width=1600,
                    height=900,
                    scale=2
                )
        except Exception:
            logging.exception(
                f"Error exportando tabla {section}"
            )

        if img_bytes:
            image_bytes_for_png.append(img_bytes)

        if table_bytes:
            image_bytes_for_png.append(table_bytes)

        section_note = (
            ""
            if (img_bytes or table_bytes)
            else "\nNota: no se generó ninguna imagen o tabla para esta sección."
        )

        report_sections.append({
            "title": section_title(section),
            "text": truncate_to_n_words(
                section_texts.get(section, ""),
                500
            ) + section_note,
            "img": img_bytes,
            "table_img": table_bytes,
            "caption": (
                f"Figura: {section_title(section)} "
                f"con los filtros seleccionados."
            ),
            "table_caption": (
                f"Tabla: {section_title(section)} "
                f"con los filtros seleccionados."
            )
        })

    if not report_sections:
        logging.warning(
            "No se generó ninguna sección para el informe."
        )
        return no_update

    if not any(
        item["text"].strip()
        for item in report_sections
    ):
        for item in report_sections:
            item["text"] = (
                f"Informe de la sección "
                f"{item['title']} "
                f"generado automáticamente."
            )

    logo_bytes = (
        base64.b64decode(LOGO_BASE64)
        if LOGO_BASE64
        else None
    )

    if (download_format or "pdf").lower() == "png":

        if not image_bytes_for_png:
            logging.warning(
                "No hay imágenes para combinar en PNG."
            )
            return no_update

        combined_png = combine_image_bytes_vertically(
            image_bytes_for_png
        )

        if not combined_png:
            return no_update

        filename = (
            f"{title.replace(' ', '_')}_"
            f"{fecha_text.replace('/', '-')}.png"
        )

        return dcc.send_bytes(
            lambda buf: buf.write(combined_png),
            filename
        )

    try:
        pdf_bytes = build_report_pdf_multi(
            title=title,
            author=author,
            logo_bytes=logo_bytes,
            sections=report_sections,
            fecha_text=fecha_text,
            filters_text=filters_text
        )

    except Exception:
        logging.exception(
            "Error generando PDF con ReportLab"
        )
        return no_update

    if not pdf_bytes:
        return no_update

    filename = (
        f"{title.replace(' ', '_')}_"
        f"{fecha_text.replace('/', '-')}.pdf"
    )

    return dcc.send_bytes(
        lambda buf: buf.write(pdf_bytes),
        filename
    )
    
@app.callback(
    Output("radar_chart", "figure"),
    [
        Input("jugador_1", "value"),
        Input("jugador_2", "value"),
        Input("game_tags", "value"),
        Input("period_tags", "value")
    ]
)
def actualizar_radar(jugador_1, jugador_2, game_tags, period_tags):
    dff_filtrado = df.copy()

    if game_tags:
        dff_filtrado = dff_filtrado[dff_filtrado["Game Tags"] == game_tags]
    if period_tags:
        dff_filtrado = dff_filtrado[dff_filtrado["Period Tags"] == period_tags]

    if not jugador_1 or not jugador_2:
        return go.Figure()

    jugadores = [jugador_1, jugador_2]
    dff_jugadores = dff_filtrado[dff_filtrado["Player Name"].isin(jugadores)]

    if dff_jugadores.empty:
        return go.Figure()

    # Agrupar métricas
    radar_data = (
        dff_jugadores.groupby("Player Name")[metricas_radar]
        .mean()
        .reset_index()
    )

    # Normalización por columna (min-max a 0–1)
    radar_data_norm = radar_data.copy()
    for m in metricas_radar:
        col_min = radar_data[m].min()
        col_max = radar_data[m].max()
        if col_max > col_min:  # evita división por cero
            radar_data_norm[m] = (radar_data[m] - col_min) / (col_max - col_min)

    # Crear figura
    fig = go.Figure()
    colores = ["#48f788", "#89bcef"]

    for idx, row in radar_data_norm.iterrows():
        hex_color = colores[idx % len(colores)]
        rgb = tuple(int(hex_color[1+i:3+i], 16) for i in (0, 2, 4))
        fill_rgba = f"rgba({rgb[0]},{rgb[1]},{rgb[2]},0.25)"

        fig.add_trace(go.Scatterpolar(
            r=row[metricas_radar].values.flatten().tolist(),
            theta=metricas_radar,
            fill="toself",
            name=row["Player Name"],
            mode="markers+lines",
            marker=dict(size=6, color=hex_color),
            line=dict(color=hex_color, width=2),
            fillcolor=fill_rgba,
            text=[f"{val:.2f}" for val in row[metricas_radar].values.flatten().tolist()],
            textposition="top center",
            textfont=dict(size=10, color="#edf1f2")
        ))

    # Estilo
    fig.update_layout(
        polar=dict(
            radialaxis=dict(
                visible=True,
                showline=True,
                linewidth=1,
                gridcolor="rgba(200,200,200,0.25)",
                gridwidth=0.8,
                tickfont=dict(size=12, color="#edf1f2")
            ),
            angularaxis=dict(
                tickfont=dict(size=12, color="#edf1f2")
            )
        ),
        showlegend=True,
        template="plotly_dark",
        title=dict(
            text=f"{jugador_1}  ||  {jugador_2}",  
        font=dict(size=20, color="#a3e3d0", family="Manrope Light"),
        x=0.5
        ),
        legend=dict(
            font=dict(size=13, color="#edf1f2"),
            orientation="h",
            yanchor="bottom",
            y=-0.25,
            xanchor="center",
            x=0.5,
            bgcolor="rgba(0,0,0,0.4)",
            bordercolor="rgba(137,188,239,0.25)",
            borderwidth=1
        )
    )

    return fig


def crear_figura(dff, metricas, referencia):
    metricas = metricas or ["Distance"]
    metricas = [m for m in metricas if m in dff.columns]
    if referencia not in dff.columns or not metricas:
        return go.Figure()
    return build_comparativas(dff, [], metricas, referencia)


def _filter_graph_dataframe(categoria, jugador, athlete, gametags, periodtags):
    dff = df.copy()
    if categoria:
        dff = dff[dff["Category"].isin(categoria)]
    if jugador:
        dff = dff[dff["Player Name"].isin(jugador)]
    if athlete:
        dff = dff[dff["Athlete Tags"].isin(athlete)]
    if gametags:
        dff = dff[dff["Game Tags"].isin(gametags)]
    if periodtags:
        dff = dff[dff["Period Tags"].isin(periodtags)]
    return dff


def _create_tab_graph_figure(
    tab,
    dff,
    categorias,
    metricas,
    referencia,
    fecha_actividad=None,
    jugador_1=None,
    jugador_2=None,
    game_tags=None,
    period_tags=None
):
    metricas = metricas or ["Distance"]
    metricas = [m for m in metricas if m in dff.columns]
    referencia = referencia or "Category"
    if tab == "comparativas":
        
        return build_comparativas(dff, categorias or [], metricas, referencia)

    if tab == "cronologico":
        title_text = build_chart_title(tab, categorias, metricas, referencia)
        cronologico = pd.melt(
                dff,
                id_vars=["Date", "Category"],
                value_vars=metricas,
                var_name="Métrica",
                value_name="Valor"
            ) 
        fig = px.scatter(
            cronologico,
            x="Date",
            y="Valor",
            color="Category",
            symbol="Métrica",
            color_discrete_sequence=["#edf1f2", "#f1a3fd", "#a3e3d0", "#89bcef", "#48f788", "#f96e83"],
            template="plotly_dark"
        )
        fig.update_traces(
            marker=dict(size=10, line=dict(width=1, color="#ffffff")),
            selector=dict(mode="markers"),
            hoverlabel=dict(bgcolor="#011c24", font_size=12, font_color="#f5f5f5"),
        )

        fig.update_layout(
            title={
                "text": title_text,  # 👈 ahora sí existe
                "font": {
                    "color": "#f5f5f5",
                    "family": "'Clash Display Semibold', 'Helvetica Neue'",
                    "size": 22
                }
            },
            paper_bgcolor="#0b0c0e",
            plot_bgcolor="#0b0c0e",
            font={"color": "#f5f5f5"},
            legend=dict(
                bgcolor="rgba(11,12,14,0.75)",
                bordercolor="#89bcef",
                borderwidth=1
            )
        )
        if LOGO_BASE64:
            fig.add_layout_image(
                dict(
                    source="data:image/png;base64," + LOGO_BASE64,
                    xref="paper",
                    yref="paper",
                    x=0.99,
                    y=0.01,
                    xanchor="right",
                    yanchor="bottom",
                    sizex=0.12,
                    sizey=0.10,
                    opacity=0.7,
                    layer="above",
                )
            )
        fig.update_xaxes(
            tickformat="%d/%m/%Y",
            showgrid=True,
            gridcolor="rgba(137,188,239,0.18)",
            zerolinecolor="rgba(255,255,255,0.08)",
            linecolor="#89bcef",
            tickfont_color="#f5f5f5",
            title_font_color="#a3e3d0",
        )
        fig.update_yaxes(
            showgrid=True,
            gridcolor="rgba(137,188,239,0.18)",
            zerolinecolor="rgba(255,255,255,0.08)",
            linecolor="#89bcef",
            tickfont_color="#f5f5f5",
            title_font_color="#a3e3d0",
        )
        return fig

    if tab == "actividad_promedios":
        if not fecha_actividad:
            return None
        fecha_dt = pd.to_datetime(fecha_actividad).normalize()
        return build_actividad_promedios_report_fig(dff, fecha_dt)

    if tab == "plyr_vs_plyr":
        return build_plyr_vs_plyr(
            dff,
            jugador_1,
            jugador_2,
            game_tag=game_tags,
            period_tag=period_tags,
            metricas=metricas,
        )

    return None


def _get_tab_filename(tab, referencia):
    if tab == "comparativas":
        referencia_label = referencia if isinstance(referencia, str) else "Grafico"
        return f"comparativas_{referencia_label.replace(' ', '_')}"
    if tab == "plyr_vs_plyr":
        return "plyr_vs_plyr"
    return tab.replace("_", "-")


def _build_graph_download(fig, filename, format):
    if fig is None or not getattr(fig, "data", None):
        logging.warning("Figura vacía o sin datos para descarga de gráfico")
        return None

    if format == "png":
        content = fig_to_png_bytes(fig, width=1600, height=900, scale=2)
    elif format == "pdf":
        png_bytes = fig_to_png_bytes(fig, width=1600, height=900, scale=2)
        if png_bytes:
            content = build_graph_pdf_bytes(filename, png_bytes)
        else:
            content = None
    else:
        logging.warning("Formato %s no soportado", format)
        return None

    if content is None:
        logging.warning("No se pudo generar el archivo %s para descarga", filename)
        return None

    return dcc.send_bytes(lambda buf: buf.write(content), f"{filename}.{format}")

@app.callback(
    Output("download-graph", "data"),
    Input("download-graph-png", "n_clicks"),
    State("tabs", "value"),
    State("categoria", "value"),
    State("jugador", "value"),
    State("athlete", "value"),
    State("gametag", "value"),
    State("periodtag", "value"),
    State("fecha-actividad", "date"),
    State("metrica", "value"),
    State("referencia", "value"),
    State("jugador_1", "value"),
    State("jugador_2", "value"),
    State("game_tags", "value"),
    State("period_tags", "value"),
    prevent_initial_call=True
)
def descargar_grafico(
    n_clicks_png,
    tab,
    categoria,
    jugador,
    athlete,
    gametags,
    periodtags,
    fecha_actividad,
    metricas,
    referencia,
    jugador_1,
    jugador_2,
    game_tags,
    period_tags,
):
    if not n_clicks_png:
        return no_update

    dff = _filter_graph_dataframe(categoria, jugador, athlete, gametags, periodtags)
    metricas = metricas or ["Distance"]
    referencia = referencia or "Category"

    fig = _create_tab_graph_figure(
        tab,
        dff,
        categoria,
        metricas,
        referencia,
        fecha_actividad=fecha_actividad,
        jugador_1=jugador_1,
        jugador_2=jugador_2,
        game_tags=game_tags,
        period_tags=period_tags,
    )

    if fig is None or not getattr(fig, "data", None):
        logging.warning("No hay figura disponible para la pestaña seleccionada: %s", tab)
        return no_update

    filename_base = _get_tab_filename(tab, referencia)
    filename = f"{filename_base}.png"
    return _build_graph_download(fig, filename, "png") or no_update



def _calc_table_height(df_export, base=400, row_height=20):
    """Calcula altura dinámica para exportar tablas según número de filas."""
    try:
        filas = len(df_export) if hasattr(df_export, "__len__") else 0
        return max(base, 40 + row_height * filas)
    except Exception:
        return base

@app.callback(
    Output("download-table", "data"),
    Input("download-table-png", "n_clicks"),
    Input("download-table-pdf", "n_clicks"),
    Input("download-table-csv", "n_clicks"),
    Input("download-table-xlsx", "n_clicks"),
    State("tabs", "value"),
    State("categoria", "value"),
    State("metrica", "value"),
    State("referencia", "value"),
    State("jugador", "value"),
    State("athlete", "value"),
    State("gametag", "value"),
    State("periodtag", "value"),
    State("fecha-actividad", "date"),
    prevent_initial_call=True
)
def descargar_tabla(
    _n_png, _n_pdf, _n_csv, _n_xlsx,
    tab, categorias, metricas, referencia,
    jugadores, athlete, gametags, periodtags, fecha_actividad
):
    ctx = dash.callback_context
    if not ctx.triggered:
        return no_update
    trigger_id = ctx.triggered[0]["prop_id"].split(".")[0]
    
    # Normalización de filtros

    if categorias and isinstance(categorias, str):
        categorias = [categorias]

    if jugadores and isinstance(jugadores, str):
        jugadores = [jugadores]

    if athlete and isinstance(athlete, str):
        athlete = [athlete]

    if gametags and isinstance(gametags, str):
        gametags = [gametags]

    if periodtags and isinstance(periodtags, str):
        periodtags = [periodtags]

    metricas = metricas or ["Distance"]

    if isinstance(metricas, str):
        metricas = [metricas]

    # --- Construir dataframe filtrado ---
    dff = df.copy()
    if categorias: dff = dff[dff["Category"].isin(categorias)]
    if jugadores: dff = dff[dff["Player Name"].isin(jugadores)]
    if athlete: dff = dff[dff["Athlete Tags"].isin(athlete)]
    if gametags: dff = dff[dff["Game Tags"].isin(gametags)]
    if periodtags: dff = dff[dff["Period Tags"].isin(periodtags)]

    metricas = metricas or ["Distance"]
    referencia = referencia or "Category"
    item_name, metadata, printed_at = build_download_metadata(tab, categorias, metricas, referencia)
    tab_name = tab_titles.get(tab, tab)

    # --- Construir df_export según tab ---
    if tab == "comparativas":
        df_export = dff.groupby(referencia)[metricas].mean().reset_index()
    elif tab == "actividad":
        fecha_dt = pd.to_datetime(fecha_actividad).normalize() if fecha_actividad else dff["Date"].max().normalize()
        dff_fecha = dff[dff["Date"].dt.normalize() == fecha_dt]
        columnas_requeridas = [
            "Player Name","Accel + Decel Efforts","Accel + Decel Efforts Per Minute","Distance",
            "Player Load","Max Velocity","Meterage Per Minute","Player Load Per Minute",
            "Sprint Distance","Sprint Efforts","Sprint Dist Per Min","High Speed Distance",
            "High Speed Efforts","High Speed Distance Per Minute","Impacts"
        ]
        columnas_presentes = [c for c in columnas_requeridas if c in dff_fecha.columns]
        df_export = dff_fecha[columnas_presentes]
    elif tab == "actividad_comparativa":
        fecha_dt = pd.to_datetime(fecha_actividad).normalize() if fecha_actividad else dff["Date"].max().normalize()
        dff_fecha = dff[dff["Date"].dt.normalize() == fecha_dt]
        metricas_base = [m for m in metricas if m in dff.columns]
        resumen_fecha = dff_fecha.groupby("Player Name")[metricas_base].mean().reset_index()
        promedio_jugador = dff.groupby("Player Name")[metricas_base].mean().reset_index()
        df_export = resumen_fecha.merge(promedio_jugador, on="Player Name", how="outer", suffixes=("", "_Promedio")).fillna(0)
    elif tab == "acwr":
        metricas_acwr = ["Distance","Player Load","Acceleration Efforts","Sprint Distance",
                         "High Speed Distance","Sprint Efforts","High Speed Efforts","Impacts"]
        dff["Player Name"] = dff["Player Name"].astype(str).str.strip()
        ultimos21 = dff["Date"].max() - pd.Timedelta(days=21)
        ultimos7 = dff["Date"].max() - pd.Timedelta(days=7)
        df21 = dff[dff["Date"] >= ultimos21]
        df7 = dff[dff["Date"] >= ultimos7]
        cronica = df21.groupby("Player Name")[metricas_acwr].mean().reset_index()
        aguda = df7.groupby("Player Name")[metricas_acwr].mean().reset_index()
        tabla = cronica.merge(aguda, on="Player Name", how="outer", suffixes=("_21", "_7"))
        tabla = tabla.loc[:, ~tabla.columns.duplicated()]
        for m in metricas_acwr:
            tabla[m + "_ACWR"] = (tabla[f"{m}_7"] / tabla[f"{m}_21"]).round(2)
        df_export = tabla[["Player Name"] + [f"{m}_ACWR" for m in metricas_acwr]].fillna(0)
    else:
            metricas_validas = [
        m
        for m in metricas
        if m in dff.columns
    ]

    if not metricas_validas:
        return no_update

    df_export = pd.melt(
        dff,
        id_vars=["Date", "Category"],
        value_vars=metricas_validas,
        var_name="Métrica",
        value_name="Valor"
    )
    # --- CSV ---
    if trigger_id == "download-table-csv":
        buffer = io.BytesIO()
        buffer.write(metadata.encode("utf-8"))
        buffer.write(b"\n")
        df_export.to_csv(buffer, index=False, line_terminator="\n")
        buffer.seek(0)
        return dcc.send_bytes(lambda b: b.write(buffer.read()), f"tabla_{tab_name}.csv")

    # --- XLSX ---
    if trigger_id == "download-table-xlsx":
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            worksheet_name = "Datos"
            df_export.to_excel(writer, sheet_name=worksheet_name, index=False, startrow=7)
            worksheet = writer.sheets[worksheet_name]
            worksheet.cell(row=1, column=1, value="Danubio Formativas 2026")
            worksheet.cell(row=2, column=1, value=f"Tabla: {item_name}")
            worksheet.cell(row=3, column=1, value=f"Impresión: {printed_at}")
            worksheet.cell(row=4, column=1, value=f"Categorías: {summarize_items(categorias, max_items=10)}")
            worksheet.cell(row=5, column=1, value=f"Métricas: {summarize_items(metricas, max_items=10)}")
            worksheet.cell(row=6, column=1, value=f"Comparar por: {referencia}")
            try:
                image = ExcelImage(str(LOGO_PATH))
                image.width = 120
                image.height = 40
                worksheet.add_image(image, "G1")
            except Exception:
                pass
        buffer.seek(0)
        return dcc.send_bytes(lambda b: b.write(buffer.read()), f"tabla_{tab_name}.xlsx")

    # --- PNG ---
    if trigger_id == "download-table-png":
        fig_table = build_section_report_table_fig(tab, dff, pd.to_datetime(fecha_actividad).normalize() if fecha_actividad else None, categorias)
        if fig_table is None or not getattr(fig_table, "data", None):
            logging.warning("No hay figura de tabla para la pestaña %s", tab)
            return no_update
        height = _calc_table_height(df_export)
        if df_export.empty:
            logging.warning(
        "No hay datos para exportar."
    )
            return no_update
        png_bytes = fig_to_png_bytes(fig_table, width=1600, height=900, scale=2)
        if not png_bytes:
            return no_update
        return dcc.send_bytes(lambda b: b.write(png_bytes), f"tabla_{tab_name}.png")

    # --- PDF ---
    if trigger_id == "download-table-pdf":
        fig_to_export = build_section_report_table_fig(tab, dff, pd.to_datetime(fecha_actividad).normalize() if fecha_actividad else None, categorias)
        if fig_to_export is None or not getattr(fig_to_export, "data", None):
            return no_update
        height = _calc_table_height(df_export, base=800)
        png_bytes = fig_to_png_bytes(fig_to_export, width=1600, height=900, scale=2)
        if not png_bytes:
            return no_update
        pdf_bytes = build_graph_pdf_bytes(f"Tabla {tab_name}", png_bytes)
        if not pdf_bytes:
            return no_update
        return dcc.send_bytes(lambda b: b.write(pdf_bytes), f"tabla_{tab_name}.pdf")

    return no_update

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 8050))
    app.run(
        host="0.0.0.0",
        port=port,
        debug=False
    )

